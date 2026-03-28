from rest_framework import generics, viewsets, status, serializers # type: ignore
from django.core.exceptions import PermissionDenied # type: ignore
from rest_framework.response import Response # type: ignore
from rest_framework.decorators import action, permission_classes as permission_classes_decorator # type: ignore
from .models import ( # type: ignore
    Trip, Expense, TravelClaim, TravelAdvance, TripOdometer, Dispute, PolicyDocument, BulkActivityBatch, JobReport,
    TravelModeMaster, BookingTypeMaster, OperatorMaster, TravelClassMaster, VehicleMaster, ProviderMaster,
    TicketStatusMaster, QuotaTypeMaster,
    LocalTravelModeMaster, LocalProviderMaster, LocalSubTypeMaster,
    StayTypeMaster, RoomTypeMaster, StayBookingTypeMaster, StayBookingSourceMaster,
    MealCategoryMaster, MealTypeMaster, MealSourceMaster, MealProviderMaster,
    IncidentalTypeMaster, CustomMasterDefinition, CustomMasterValue, MasterModule, TripTracking
)
from .serializers import ( # type: ignore
    TripSerializer, ExpenseSerializer, TravelClaimSerializer, TravelAdvanceSerializer,
    TripOdometerSerializer, DisputeSerializer, PolicyDocumentSerializer, PolicyDocumentDetailSerializer, BulkActivityBatchSerializer, JobReportSerializer,
    TravelModeMasterSerializer, BookingTypeMasterSerializer, OperatorMasterSerializer, TravelClassMasterSerializer,
    VehicleMasterSerializer, ProviderMasterSerializer, TicketStatusMasterSerializer, QuotaTypeMasterSerializer,
    LocalTravelModeMasterSerializer, LocalProviderMasterSerializer, LocalSubTypeMasterSerializer,
    StayTypeMasterSerializer, RoomTypeMasterSerializer, StayBookingTypeMasterSerializer, StayBookingSourceMasterSerializer,
    MealCategoryMasterSerializer, MealTypeMasterSerializer, MealSourceMasterSerializer, MealProviderMasterSerializer,
    IncidentalTypeMasterSerializer, CustomMasterDefinitionSerializer, CustomMasterValueSerializer, MasterModuleSerializer,
    TripTrackingSerializer
)
import io # type: ignore
import json # type: ignore
import pandas as pd # type: ignore
from django.http import HttpResponse # type: ignore
from rest_framework.permissions import AllowAny # type: ignore
from django.db.models import Q # type: ignore
import base64 # type: ignore
import binascii # type: ignore
from api_management.utils import encrypt_key, decrypt_key # type: ignore
from django.utils import timezone # type: ignore
from rest_framework.views import APIView # type: ignore
from django.db.models import Sum # type: ignore
from core.models import User # type: ignore
from notifications.models import Notification # type: ignore
from core.permissions import IsCustomAuthenticated # type: ignore
import requests # type: ignore
import datetime # type: ignore

def decode_id(encoded_id):
    s_id = str(encoded_id) if encoded_id else ""
    if not s_id:
        return None
        
    try:
        # Check if it's already a numeric-looking ID or doesn't look like base64
        if s_id.isdigit() or s_id.startswith('TRP-') or s_id.startswith('ITS-'):
            return s_id
            
        padding = 4 - (len(s_id) % 4)
        if padding != 4:
            s_id += '=' * padding
        
        s_id = s_id.replace('-', '+').replace('_', '/')
        
        decoded_bytes = base64.b64decode(s_id)
        return decoded_bytes.decode('utf-8')
    except (binascii.Error, UnicodeDecodeError, ValueError):
        return encoded_id

def _is_finance_head(user):
    """Checks if a user is the Finance Head."""
    user_role = (user.role.name.lower() if user.role else '')
    dept = user.department.lower()
    desig = user.designation.lower()
    return 'head' in dept and 'finance' in dept or 'head' in desig and 'finance' in desig or 'cfo' in user_role

def _is_finance_executive(user):
    """Checks if a user is a Finance Executive."""
    if _is_finance_head(user): return False
    user_role = (user.role.name.lower() if user.role else '')
    dept = user.department.lower()
    desig = user.designation.lower()
    return 'finance' in user_role or 'finance' in dept or 'finance' in desig

def _is_hr(user):
    """Checks if a user is an HR user."""
    user_role = (user.role.name.lower() if user.role else '')
    dept = user.department.lower()
    desig = user.designation.lower()
    return 'hr' in user_role or 'hr' in dept or 'hr' in desig

def _get_finance_users():
    """Returns a list of users who should be treated as Finance."""
    all_users = User.objects.filter(is_active=True).select_related('role')
    return [u for u in all_users if _is_finance_executive(u) or _is_finance_head(u)]

def _get_hr_users():
    """Returns a list of users who should be treated as HR."""
    all_users = User.objects.filter(is_active=True).select_related('role')
    return [u for u in all_users if _is_hr(u)]

def get_finance_head(user, exclude_user=None):
    """Finds a finance approver (Head of Finance)."""
    all_finance = _get_finance_users()
    heads = [u for u in all_finance if _is_finance_head(u)]
    if exclude_user:
        heads = [u for u in heads if u.id != exclude_user.id]
    
    if heads:
        return heads[0]
    # Fallback to any finance user if no head found
    return all_finance[0] if all_finance else None

def get_finance_executive(user, exclude_user=None):
    """Finds a finance executive."""
    all_finance = _get_finance_users()
    execs = [u for u in all_finance if _is_finance_executive(u)]
    if exclude_user:
         execs = [u for u in execs if u.id != exclude_user.id]
         
    if execs:
        return execs[0]
    return all_finance[0] if all_finance else None

def get_hr_head(user):
    """Finds an HR approver (Head of HR)."""
    all_hr = _get_hr_users()
    # Try local HR first
    local_hr = [u for u in all_hr if u.base_location == user.base_location]
    if local_hr:
        return local_hr[0]
    
    return all_hr[0] if all_hr else None

def notify_hr(title, message):
    """Notify all users with HR role."""
    hr_users = User.objects.filter(role__name__icontains='hr')
    for hr in hr_users:
        Notification.objects.create(
            user=hr,
            title=title,
            message=message,
            type='info'
        )


def update_trip_lifecycle(trip, title, description):
    """Helper to append events to the Trip's JSON lifecycle field."""
    if not trip:
        return
    event = {
        "title": title,
        "status": "completed",
        "date": timezone.now().strftime("%b %d, %Y"),
        "description": description
    }
    events = trip.lifecycle_events
    if not isinstance(events, list):
        events = []
    
    # Avoid duplicate events with same title and description
    if not any(e.get('title') == title and e.get('description') == description for e in events):
        events.append(event)
        trip.lifecycle_events = events
        trip.save(update_fields=['lifecycle_events'])






def resolve_approver(user, members_data=None):
    """Helper to resolve the first approver in the management hierarchy."""
    def is_admin(u):
        if not u or not u.role:
            return False
        return u.role.name.lower() in ['admin', 'it-admin', 'superuser', 'it admin', 'system administrator']
    
    reporting_manager = user.reporting_manager
    senior_manager = user.senior_manager
    hod_director = user.hod_director
    
    current_approver = reporting_manager if not is_admin(reporting_manager) else None
    h_level = 1
    
    if not current_approver:
        current_approver = senior_manager if not is_admin(senior_manager) else None
        h_level = 2
    
    if not current_approver:
        current_approver = hod_director if not is_admin(hod_director) else None
        h_level = 3
        
    if not current_approver:
        # Fallback to members' managers if applicable
        potential_managers = []
        if members_data:
            import re
            for m_str in members_data:
                match = re.search(r'\((.*?)\)', m_str)
                if match:
                    member_id = match.group(1)
                    member_user = User._get_or_create_shell_user(member_id)
                    manager = member_user.reporting_manager if member_user else None
                    if manager and not is_admin(manager):
                        potential_managers.append(manager)
            
        if potential_managers:
            potential_managers.sort(key=lambda m: getattr(m, 'level_rank', 99))
            current_approver = potential_managers[0]
            h_level = 1
        else:
            current_approver = get_hr_head(user)
            h_level = 1
            
    return current_approver, h_level, reporting_manager, senior_manager, hod_director

class TripListCreateView(generics.ListCreateAPIView):
    serializer_class = TripSerializer
    permission_classes = [IsCustomAuthenticated]

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return Trip.objects.none()
            
        all_trips = self.request.query_params.get('all') == 'true'
        user_role = user.role.name.lower() if user.role else ''
        
        if all_trips:
            if user_role in ['admin', 'guesthousemanager', 'finance', 'cfo']:
                return Trip.objects.all().order_by('-created_at')
            
            return Trip.objects.filter(
                Q(user=user) | 
                Q(current_approver=user)
            ).distinct().order_by('-created_at')
            
        queryset = Trip.objects.filter(user=user, consider_as_local=False)
        search_query = self.request.query_params.get('search')
        if search_query:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(trip_id__icontains=search_query) |
                Q(purpose__icontains=search_query) |
                Q(source__icontains=search_query) |
                Q(destination__icontains=search_query)
            )
        return queryset.order_by('-created_at')

    def perform_create(self, serializer, is_local=False):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            from rest_framework.exceptions import AuthenticationFailed # type: ignore
            raise AuthenticationFailed("Authentication required.")
        
        # Admin / Superuser skip approvals
        user_role = user.role.name.lower() if user.role else ''
        if user_role in ['admin', 'superuser', 'it-admin']:
            trip = serializer.save(
                user=user,
                status='Approved',
                current_approver=None,
                hierarchy_level=0,
                consider_as_local=is_local
            )
            label = "Travel" if is_local else "Trip"
            update_trip_lifecycle(trip, "Auto-Approved", f"{label} request auto-approved for Administrator.")
            return

        members_data = serializer.validated_data.get('members', [])
        current_approver, h_level, rm, sm, hod = resolve_approver(user, members_data)
        
        try:
            trip = serializer.save(
                user=user,
                status='Pending',
                current_approver=current_approver,
                hierarchy_level=h_level,
                consider_as_local=is_local,
                # Snapshots
                user_name=user.name,
                user_designation=user.designation,
                user_department=user.department,
                reporting_manager_name=rm.name if rm else None,
                senior_manager_name=sm.name if sm else None,
                hod_director_name=hod.name if hod else None
            )
        except Exception as e:
            # convert DB errors to validation error so frontend sees message
            # also log full traceback on server for diagnostics
            import traceback
            traceback.print_exc()
            from rest_framework.exceptions import ValidationError # type: ignore # type: ignore # type: ignore
            msg = str(e)
            print("ERROR IN TRIP CREATION:", msg)
            raise ValidationError({"detail": msg})

        label = "Travel" if is_local else "Trip"
        if current_approver:
            Notification.objects.create(
                user=current_approver,
                title=f"New {label} Request",
                message=f"{user.name} has submitted a new {label.lower()} request to {trip.destination}.",
                type='info'
            )
        
        if trip.accommodation_requests and any('Room' in r for r in trip.accommodation_requests):
            gh_managers = User.objects.filter(role__name='GuestHouseManager', is_active=True)
            for manager in gh_managers:
                Notification.objects.create(
                    user=manager,
                    title="Room Request Received",
                    message=f"{user.name} has requested a room for {label.lower()} {trip.trip_id}.",
                    type='info'
                )

        notify_hr(f"New {label} Request", f"{user.name} has raised a {label.lower()} request to {trip.destination} (ID: {trip.trip_id}).")

class TravelListCreateView(TripListCreateView):
    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user: return Trip.objects.none()
        queryset = Trip.objects.filter(user=user, consider_as_local=True)
        search_query = self.request.query_params.get('search')
        if search_query:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(trip_id__icontains=search_query) |
                Q(purpose__icontains=search_query) |
                Q(source__icontains=search_query) |
                Q(destination__icontains=search_query)
            )
        return queryset.order_by('-created_at')

    def perform_create(self, serializer, is_local=True): # type: ignore
        super().perform_create(serializer, is_local=is_local)

class TripBookingSearchView(generics.ListAPIView):
    serializer_class = TripSerializer
    permission_classes = [IsCustomAuthenticated] 

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return Trip.objects.none()

        # Users only search their own trips in this view
        queryset = Trip.objects.filter(user=user)

        search_query = self.request.query_params.get('search', None)
        if search_query:
            try:
                import base64
                padding = len(search_query) % 4
                if padding: 
                    search_query += '=' * (4 - padding)
                search_query = base64.b64decode(search_query).decode('utf-8')
            except Exception as e:
                print(f"Decoding error: {e}")
                pass

            queryset = queryset.filter(
                Q(trip_id__istartswith=search_query) | 
                Q(purpose__istartswith=search_query) | 
                Q(source__istartswith=search_query) | 
                Q(destination__istartswith=search_query) | 
                Q(trip_leader__istartswith=search_query)
            )
        
        return queryset

class TripDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Trip.objects.all()
    serializer_class = TripSerializer
    permission_classes = [IsCustomAuthenticated]
    lookup_field = 'trip_id'

    def get_object(self):
        if 'trip_id' in self.kwargs:
            self.kwargs['trip_id'] = decode_id(self.kwargs['trip_id'])
        
        obj = super().get_object()
        user = getattr(self.request, 'custom_user', None)
        if not user:
            raise PermissionDenied("Not authenticated")
            
        user_role = (user.role.name.lower() if user.role else '')
        is_admin = user_role in ['admin', 'it-admin', 'superuser']
        is_finance = user_role in ['finance', 'cfo']
        is_gh_manager = user_role == 'guesthousemanager'
        
        # Authorization check: Owner, Managers in the hierarchy, Current Approver, Finance, Guest House Manager, or Admin
        is_owner = (obj.user == user)
        # Check if the current user is any of the requester's managers or the current approver
        is_manager = user in [obj.user.reporting_manager, obj.user.senior_manager, obj.user.hod_director, obj.current_approver]
        
        if not (is_owner or is_manager or is_admin or is_finance or is_gh_manager):
             self.permission_denied(self.request, message="Not authorized to view this trip details")
             
        return obj

class TripTrackingView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request, trip_id):
        print(f"DEBUG: TripTrackingView.get called for trip_id: {trip_id}")
        real_trip_id = decode_id(trip_id)
        # Verify trip exists and user has access
        try:
            trip = Trip.objects.get(trip_id=real_trip_id)
        except Trip.DoesNotExist:
            print(f"DEBUG: Trip {real_trip_id} not found")
            return Response({"error": "Trip not found"}, status=status.HTTP_404_NOT_FOUND)

        # Basic access check: requester or manager or finance or admin
        user = getattr(request, 'custom_user', None)
        print(f"DEBUG: Requester: {user.employee_id if user else 'Anonymous'}")
        
        # ... existing logic ...
        is_owner = (trip.user == user)
        is_manager = False
        if user:
            is_manager = user in [trip.user.reporting_manager, trip.user.senior_manager, trip.user.hod_director, trip.current_approver]
        
        user_role = user.role.name.lower() if user and user.role else ''
        is_privileged = user_role in ['admin', 'finance', 'cfo', 'guesthousemanager']

        if not (is_owner or is_manager or is_privileged):
            print(f"DEBUG: Unauthorized access attempt to trip {real_trip_id}")
            return Response({"error": "Unauthorized"}, status=status.HTTP_403_FORBIDDEN)

        tracking_data = TripTracking.objects.filter(trip=trip).order_by('timestamp')
        print(f"DEBUG: Returning {tracking_data.count()} points")
        serializer = TripTrackingSerializer(tracking_data, many=True)
        return Response(serializer.data)

    def post(self, request, trip_id):
        print(f"DEBUG: TripTrackingView.post called for trip_id: {trip_id}")
        real_trip_id = decode_id(trip_id)
        print(f"DEBUG: Decoded trip_id: {real_trip_id}")
        
        trip = None
        try:
            trip = Trip.objects.get(trip_id=real_trip_id)
        except Trip.DoesNotExist:
            # Fallback: check if the trip exists but is soft-deleted
            try:
                deleted_trip = Trip.all_objects.get(trip_id=real_trip_id)
                print(f"DEBUG: Trip {real_trip_id} exists but is soft-deleted (is_deleted={deleted_trip.is_deleted}). Allowing tracking.")
                trip = deleted_trip  # Allow tracking even for soft-deleted trips
            except Trip.DoesNotExist:
                print(f"DEBUG: Trip {real_trip_id} does NOT exist in DB at all (even in all_objects). POST rejected.")
                return Response({"error": "Trip not found"}, status=status.HTTP_404_NOT_FOUND)

        # Only trip owner can post tracking points
        user = getattr(request, 'custom_user', None)
        print(f"DEBUG: POST Requester: {user.employee_id if user else 'Anonymous'}")
        
        if not user or trip.user != user:
            print(f"DEBUG: POST Unauthorized for user {user.employee_id if user else 'None'}")
            return Response({"error": "Only trip owner can submit tracking data"}, status=status.HTTP_403_FORBIDDEN)

        data = request.data.copy()
        data['trip'] = trip.trip_id
        
        serializer = TripTrackingSerializer(data=data)
        if serializer.is_valid():
            tracking = serializer.save()
            print("DEBUG: Tracking point saved successfully")
            
            try:
                from django.db import connection
                if 'travel_tripgeofencelocationset' in connection.introspection.table_names():
                    from .models import TripGeofenceLocationSet
                    geofence_set, created = TripGeofenceLocationSet.objects.get_or_create(trip=trip)
                    loc_data = geofence_set.location_data
                    if not isinstance(loc_data, list):
                        loc_data = []
                    # add point to the set
                    loc_data.append({
                        "latitude": float(tracking.latitude),
                        "longitude": float(tracking.longitude),
                        "timestamp": tracking.timestamp.isoformat() if tracking.timestamp else None,
                        "accuracy": tracking.accuracy,
                        "speed": tracking.speed
                    })
                    geofence_set.location_data = loc_data
                    geofence_set.last_latitude = tracking.latitude
                    geofence_set.last_longitude = tracking.longitude
                    geofence_set.save()
            except Exception as e:
                print(f"DEBUG: Error updating Geofence Location Set: {e}")
            # --------------------------------------

            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        print(f"DEBUG: Serializer errors: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class TeamLiveTrackingView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        user = getattr(request, 'custom_user', None)
        if not user:
            return Response({"error": "Unauthorized"}, status=403)
        
        from django.utils import timezone
        today = timezone.now().date()
        
        # Get ALL active approved trips today (start_date <= today <= end_date)
        # We must filter reporting_manager in Python since it's a dynamic property on the User model
        active_trips = Trip.all_objects.filter(
            start_date__lte=today,
            end_date__gte=today,
            status__iexact='Approved'
        )

        from core.models import Session
        from django.db import connection
        
        has_geofence_table = 'travel_tripgeofencelocationset' in connection.introspection.table_names()
        if has_geofence_table:
            from travel.models import TripGeofenceLocationSet

        results = []
        for trip in active_trips:
            if not trip.user:
                continue
                
            # Evaluate the property to check if current user is the manager
            rm = trip.user.reporting_manager
            if not rm or rm.employee_id != user.employee_id:
                continue

            # Check if the employee is currently logged out based on their most recent session
            latest_session = Session.objects.filter(user=trip.user).order_by('-created_at').first()
            is_logged_in = latest_session.is_active if latest_session else False

            # In case the user is logged out, show the last synced geofence from TripGeofenceLocationSet
            geofence_set = None
            if has_geofence_table:
                try:
                    geofence_set = TripGeofenceLocationSet.objects.filter(trip=trip).first()
                except Exception:
                    pass
            
            latest_tracking = TripTracking.objects.filter(trip=trip).order_by('-timestamp').first()
            
            lat = float(latest_tracking.latitude) if latest_tracking else (float(geofence_set.last_latitude) if geofence_set and geofence_set.last_latitude else None)
            lng = float(latest_tracking.longitude) if latest_tracking else (float(geofence_set.last_longitude) if geofence_set and geofence_set.last_longitude else None)
            accuracy = latest_tracking.accuracy if latest_tracking else None
            speed = latest_tracking.speed if latest_tracking else None
            last_updated = latest_tracking.timestamp if latest_tracking else (geofence_set.last_updated if geofence_set else None)

            results.append({
                "trip_id": trip.trip_id,
                "employee_name": trip.user.name,
                "employee_id": trip.user.employee_id,
                "destination": trip.destination,
                "purpose": trip.purpose,
                "consider_as_local": trip.consider_as_local,
                "status": trip.status,
                "is_logged_out": not is_logged_in,
                "latitude": lat,
                "longitude": lng,
                "last_updated": last_updated,
                "accuracy": accuracy,
                "speed": speed,
            })

        return Response(results, status=200)

class ApprovalCountView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        user = getattr(request, 'custom_user', None)
        if not user:
            return Response({"count": 0})
        
        user_role = (user.role.name.lower() if user.role else '')
        is_admin = user_role in ['admin', 'it-admin', 'superuser']
        is_finance = 'finance' in user_role
        
        if is_admin:
            trip_count = Trip.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded']).count()
            advance_count = TravelAdvance.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded']).count()
            claim_count = TravelClaim.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded']).count()
        elif is_finance:
            if user.office_level == 1:
                # Finance Head counts
                trip_count = 0
                advance_count = TravelAdvance.objects.filter(status='PENDING_HEAD').count()
                claim_count = TravelClaim.objects.filter(status='PENDING_HEAD').count()
            else:
                # Finance Executive counts
                trip_count = 0
                pending_money_statuses = ['PENDING_EXECUTIVE', 'HR Approved', 'REJECTED_BY_HEAD', 'PENDING_FINAL_RELEASE', 'Approved', 'Under Process']
                advance_count = TravelAdvance.objects.filter(status__in=pending_money_statuses).count()
                claim_count = TravelClaim.objects.filter(status__in=pending_money_statuses).count()
        else:
            # Manager counts
            trip_count = Trip.objects.filter(current_approver=user, status__in=['Pending', 'Submitted', 'Forwarded', 'Manager Approved']).count()
            advance_count = TravelAdvance.objects.filter(current_approver=user, status__in=['Pending', 'Submitted', 'Forwarded', 'Manager Approved']).count()
            claim_count = TravelClaim.objects.filter(current_approver=user, status__in=['Pending', 'Submitted', 'Forwarded', 'Manager Approved']).count()
            
        return Response({
            "total": trip_count + advance_count + claim_count,
            "trips": trip_count,
            "advances": advance_count,
            "claims": claim_count
        })


class ApprovalsView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        user = getattr(request, 'custom_user', None)
        if not user:
            return Response({"error": "User not found"}, status=401)
        
        user_role = (user.role.name.lower() if user.role else '')
        is_admin = user_role in ['admin', 'it-admin', 'superuser']
        is_finance = _is_finance_executive(user) or _is_finance_head(user)
        is_hr = _is_hr(user)
        is_finance_head = _is_finance_head(user)
        
        tab = request.query_params.get('tab', 'pending')
        type_filter = request.query_params.get('type', 'all') 
        
        trips = Trip.objects.none()
        advances = TravelAdvance.objects.none()
        claims = TravelClaim.objects.none()
        disputes = Dispute.objects.none()

        if tab == 'history':
            from core.models import AuditLog # type: ignore
            # Include 'UPDATE' to capture older approvals or edits made by managers
            involved_logs = AuditLog.objects.filter(user=user, action__in=['APPROVE', 'FORWARD', 'REJECT', 'UPDATE'])
            
            trip_pks = involved_logs.filter(model_name='Trip').values_list('object_id', flat=True)
            advance_pks_raw = involved_logs.filter(model_name='TravelAdvance').values_list('object_id', flat=True)
            claim_pks_raw = involved_logs.filter(model_name='TravelClaim').values_list('object_id', flat=True)
            
            # Convert string IDs to integers for numeric primary keys
            advance_pks = [int(pk) for pk in advance_pks_raw if pk and pk.isdigit()]
            claim_pks = [int(pk) for pk in claim_pks_raw if pk and pk.isdigit()]
            
            trips = Trip.objects.filter(trip_id__in=trip_pks)
            advances = TravelAdvance.objects.filter(id__in=advance_pks)
            claims = TravelClaim.objects.filter(id__in=claim_pks)
            
            # Admins see everything in history
            if is_admin:
                history_statuses = ['Approved', 'Rejected', 'Resolved', 'Paid', 'HR Approved', 'Manager Approved', 'COMPLETED']
                trips = Trip.objects.filter(status__in=history_statuses)
                advances = TravelAdvance.objects.filter(status__in=history_statuses)
                claims = TravelClaim.objects.filter(status__in=history_statuses)
        else:
            # Multi-Tab Workflow Logic (Action Required, Under Process, etc.)
            if tab == 'processing':
                trips = Trip.objects.filter(status='Under Process')
                advances = TravelAdvance.objects.filter(status='Under Process')
                claims = TravelClaim.objects.filter(status='Under Process')
                # For Managers, also filter by current_approver
                if not is_finance and not is_admin:
                    trips = trips.filter(current_approver=user)
                    advances = advances.filter(current_approver=user)
                    claims = claims.filter(current_approver=user)
            elif tab == 'completed':
                completed_statuses = ['Paid', 'COMPLETED', 'Settled']
                trips = Trip.objects.filter(status__in=completed_statuses)
                advances = TravelAdvance.objects.filter(status__in=completed_statuses)
                claims = TravelClaim.objects.filter(status__in=completed_statuses)
                if not is_finance and not is_admin:
                    trips = trips.filter(user=user)
                    advances = advances.filter(trip__user=user)
                    claims = claims.filter(trip__user=user)
            elif tab == 'rejected':
                rejected_statuses = ['Rejected', 'Rejected by Finance', 'Cancelled']
                trips = Trip.objects.filter(status__in=rejected_statuses)
                advances = TravelAdvance.objects.filter(status__in=rejected_statuses)
                claims = TravelClaim.objects.filter(status__in=rejected_statuses)
                if not is_finance and not is_admin:
                    trips = trips.filter(user=user)
                    advances = advances.filter(trip__user=user)
                    claims = claims.filter(trip__user=user)
            else:
                # DEFAULT: Action Required (Pending)
                if is_admin:
                    finance_pending = ['PENDING_EXECUTIVE', 'PENDING_HEAD', 'PENDING_FINAL_RELEASE', 'REJECTED_BY_HEAD']
                    trips = Trip.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Manager Approved', 'HR Approved'] + finance_pending)
                    advances = TravelAdvance.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Manager Approved', 'HR Approved'] + finance_pending)
                    claims = TravelClaim.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Manager Approved', 'HR Approved'] + finance_pending)
                elif is_finance:
                    if is_finance_head:
                        advances = TravelAdvance.objects.filter(status='PENDING_HEAD')
                        claims = TravelClaim.objects.filter(status='PENDING_HEAD')
                    else:
                        pending_money_statuses = ['PENDING_EXECUTIVE', 'REJECTED_BY_HEAD', 'PENDING_FINAL_RELEASE', 'HR Approved', 'Approved']
                        advances = TravelAdvance.objects.filter(status__in=pending_money_statuses)
                        claims = TravelClaim.objects.filter(status__in=pending_money_statuses)
                    
                    # Filter by project/dept if needed
                    finance_dept = user.department
                    if finance_dept and finance_dept.lower() not in ['finance', 'finance department', 'accounts', 'finance head dept', 'finance executive dept']:
                        advances = advances.filter(trip__project_code__istartswith=finance_dept)
                        claims = claims.filter(trip__project_code__istartswith=finance_dept)
                elif is_hr:
                    # HR verification stage
                    trips = Trip.objects.filter(status='Manager Approved')
                    advances = TravelAdvance.objects.filter(status='Manager Approved')
                    claims = TravelClaim.objects.filter(status='Manager Approved')
                else:
                    # Regular hierarchy
                    trips = Trip.objects.filter(current_approver=user, status__in=['Pending', 'Submitted', 'Forwarded'])
                    advances = TravelAdvance.objects.filter(current_approver=user, status__in=['Pending', 'Submitted', 'Forwarded'])
                    claims = TravelClaim.objects.filter(current_approver=user, status__in=['Pending', 'Submitted', 'Forwarded'])
        
        # Apply search filter if provided
        search_query = request.query_params.get('search')
        if search_query:
            trips = trips.filter(
                Q(trip_id__icontains=search_query) |
                Q(user__name__icontains=search_query) |
                Q(purpose__icontains=search_query) |
                Q(source__icontains=search_query) |
                Q(destination__icontains=search_query)
            ).distinct()
            advances = advances.filter(
                Q(trip__trip_id__icontains=search_query) |
                Q(trip__user__name__icontains=search_query) |
                Q(purpose__icontains=search_query)
            ).distinct()
            claims = claims.filter(
                Q(trip__trip_id__icontains=search_query) |
                Q(trip__user__name__icontains=search_query)
            ).distinct()

        tasks = []
        # Support filtering by type if specified
        if type_filter in ['all', 'trip']:
            for t in trips.order_by('-created_at'):
                tasks.append({
                    "id": f"TRIP-{t.trip_id}", "db_id": t.trip_id, "type": "Trip",
                    "requester": t.user.name if t.user else "Unknown", "purpose": t.purpose,
                    "status": t.status, "date": t.created_at.strftime("%b %d, %Y"),
                    "hierarchy_level": t.hierarchy_level,
                    "trip_id": t.trip_id,
                    "is_local": t.consider_as_local,
                    "cost": t.cost_estimate,
                    "details": {
                        "source": t.source, "destination": t.destination, 
                        "start_date": t.start_date.strftime("%b %d, %Y"),
                        "end_date": t.end_date.strftime("%b %d, %Y"),
                        "travel_mode": t.travel_mode,
                        "composition": t.composition,
                        "vehicle_type": t.vehicle_type,
                        "purpose": t.purpose,
                        "project_code": t.project_code,
                        "job_reports": [
                            {
                                "id": jr.id,
                                "created_at": jr.created_at.strftime("%b %d, %Y"),
                                "user_name": jr.user.name if jr.user else "N/A",
                                "description": jr.description,
                                "attachment": jr.attachment,
                                "file_name": jr.file_name
                            } for jr in t.job_reports.all()
                        ],
                        "odometer": {
                            "start_reading": str(t.odometer_details.start_odo_reading) if hasattr(t, 'odometer_details') and t.odometer_details.start_odo_reading else None,
                            "start_image": decrypt_key(t.odometer_details.start_odo_image) if hasattr(t, 'odometer_details') and t.odometer_details.start_odo_image else None,
                            "end_reading": str(t.odometer_details.end_odo_reading) if hasattr(t, 'odometer_details') and t.odometer_details.end_odo_reading else None,
                            "end_image": decrypt_key(t.odometer_details.end_odo_image) if hasattr(t, 'odometer_details') and t.odometer_details.end_odo_image else None,
                        } if hasattr(t, 'odometer_details') else None
                    }
                })
            
        if type_filter in ['all', 'advance']:
            for a in advances.order_by('-created_at'):
                tasks.append({
                    "id": f"ADV-{a.id}", "db_id": a.id, "type": "Money Top-up / Advance",
                    "requester": a.trip.user.name if a.trip.user else "Unknown",
                    "purpose": f"Advance: {a.purpose}", "cost": f"₹{a.requested_amount}",
                    "status": a.status, "date": a.created_at.strftime("%b %d, %Y"),
                    "hierarchy_level": a.hierarchy_level,
                    "trip_id": a.trip.trip_id,
                    "is_local": a.trip.consider_as_local,
                    "details": {
                        "source": a.trip.source,
                        "destination": a.trip.destination,
                        "requested_amount": str(a.requested_amount),
                        "hr_approved_amount": str(a.hr_approved_amount or 0),
                        "hr_remarks": a.hr_remarks or "",
                        "executive_approved_amount": str(a.executive_approved_amount),
                        "reason": a.purpose,
                        "trip_destination": a.trip.destination,
                        "trip_id": a.trip.trip_id,
                        "start_date": a.trip.start_date.strftime("%b %d, %Y") if a.trip.start_date else "N/A",
                        "end_date": a.trip.end_date.strftime("%b %d, %Y") if a.trip.end_date else "N/A",
                    }
                })

        if type_filter in ['all', 'expense', 'mileage']:
            for c in claims.order_by('-created_at'):
                tasks.append({
                    "id": f"CLAIM-{c.id}", "db_id": c.id, "type": "Expense Claim",
                    "requester": c.trip.user.name if c.trip.user else "Unknown",
                    "purpose": f"Claim for {c.trip.destination}", "cost": f"₹{c.total_amount}",
                    "status": c.status, "date": c.created_at.strftime("%b %d, %Y"),
                    "hierarchy_level": c.hierarchy_level,
                    "trip_id": c.trip.trip_id,
                    "is_local": c.trip.consider_as_local,
                    "details": {
                        "source": c.trip.source,
                        "destination": c.trip.destination,
                        "total_amount": str(c.total_amount),
                        "requested_amount": str(c.total_amount),
                        "approved_amount": str(c.approved_amount),
                        "hr_approved_amount": str(c.hr_approved_amount or 0),
                        "hr_remarks": getattr(c, "hr_remarks", ""),
                        "executive_approved_amount": str(c.executive_approved_amount),
                        "trip_id": c.trip.trip_id,
                        "start_date": c.trip.start_date.strftime("%b %d, %Y") if c.trip.start_date else "N/A",
                        "end_date": c.trip.end_date.strftime("%b %d, %Y") if c.trip.end_date else "N/A",
                        "expenses": [
                            {
                                "id": e.id,
                                "category": e.category,
                                "date": e.date.strftime("%b %d, %Y"),
                                "amount": str(e.amount),
                                "description": e.description,
                                "status": e.status,
                                "receipt_image": decrypt_key(e.receipt_image) if e.receipt_image else "",
                                "rm_remarks": e.rm_remarks or "",
                                "hr_remarks": e.hr_remarks or "",
                                "finance_remarks": e.finance_remarks or ""
                            } for e in c.trip.expenses.all()
                        ],
                        "job_reports": [
                            {
                                "id": jr.id,
                                "created_at": jr.created_at.strftime("%b %d, %Y"),
                                "user_name": jr.user.name if jr.user else "N/A",
                                "description": jr.description,
                                "attachment": jr.attachment,
                                "file_name": jr.file_name
                            } for jr in c.trip.job_reports.all()
                        ],
                        "odometer": {
                            "start_reading": str(c.trip.odometer_details.start_odo_reading) if hasattr(c.trip, 'odometer_details') and c.trip.odometer_details.start_odo_reading else None,
                            "start_image": decrypt_key(c.trip.odometer_details.start_odo_image) if hasattr(c.trip, 'odometer_details') and c.trip.odometer_details.start_odo_image else None,
                            "end_reading": str(c.trip.odometer_details.end_odo_reading) if hasattr(c.trip, 'odometer_details') and c.trip.odometer_details.end_odo_reading else None,
                            "end_image": decrypt_key(c.trip.odometer_details.end_odo_image) if hasattr(c.trip, 'odometer_details') and c.trip.odometer_details.end_odo_image else None,
                        } if hasattr(c.trip, 'odometer_details') else None
                    }
                })

        for d in disputes.order_by('-created_at'):
            tasks.append({
                "id": f"DISPUTE-{d.id}", "db_id": d.id, "type": "Dispute",
                "requester": d.raised_by.name if d.raised_by else "Unknown",
                "purpose": f"Dispute: {d.category}", "status": d.status,
                "date": d.created_at.strftime("%b %d, %Y")
            })

        return Response(tasks)

    def post(self, request):
        user = getattr(request, 'custom_user', None)
        task_id = request.data.get('id')
        action = request.data.get('action') 
        
        if not task_id or not action:
            return Response({"error": "ID and Action required"}, status=400)
            
        if action == 'UpdateItem':
            item_id = request.data.get('item_id')
            item_status = request.data.get('item_status')
            remarks = request.data.get('remarks', '')
            if item_id:
                from .models import Expense # type: ignore
                expense = Expense.objects.filter(id=item_id).first()
                if expense:
                    update_fields = {'status': item_status}
                    # Determine which remarks field to update based on user role
                    user_role = user.role.name.upper() if user and user.role else ""
                    if "FINANCE" in user_role:
                        update_fields['finance_remarks'] = remarks
                    elif "HR" in user_role:
                        update_fields['hr_remarks'] = remarks
                    else:
                        update_fields['rm_remarks'] = remarks
                    
                    Expense.objects.filter(id=item_id).update(**update_fields)
                    return Response({"message": f"Item {item_id} set to {item_status} with remarks"})
            return Response({"error": "Item ID required or not found"}, status=400)

        try:
            if task_id.startswith('TRIP-'):
                obj = Trip.objects.get(trip_id=task_id.replace('TRIP-', ''))
            elif task_id.startswith('ADV-'):
                obj = TravelAdvance.objects.get(id=task_id.replace('ADV-', ''))
            elif task_id.startswith('CLAIM-'):
                obj = TravelClaim.objects.get(id=task_id.replace('CLAIM-', ''))
            elif task_id.startswith('DISPUTE-'):
                obj = Dispute.objects.get(id=task_id.replace('DISPUTE-', ''))
            else:
                return Response({"error": "Invalid task ID"}, status=400)
            
            self._handle_workflow(obj, action, user, request.data)
            return Response({"message": f"Task processed successfully: {action}"})
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({"error": str(e)}, status=400)

    def _handle_workflow(self, obj, action, user, data=None):
        """Generic multi-level workflow handler"""
        # Determine requester 
        requester = obj.user if hasattr(obj, 'user') else obj.trip.user
        request_type = "Trip" if isinstance(obj, Trip) else ("Advance" if isinstance(obj, TravelAdvance) else "Expense Claim")
        
        # Security Check: Determine roles and check authorization early
        user_role = user.role.name.lower() if user.role else ''
        is_admin = user_role in ['admin', 'it-admin', 'superuser']
        is_hr = _is_hr(user)
        is_finance_exec = _is_finance_executive(user)
        is_finance_head = _is_finance_head(user)
        is_finance = is_finance_exec or is_finance_head

        if not is_admin:
            authorized = False
            # 1. Primary approver check
            if obj.current_approver == user:
                authorized = True
            # 2. Functional role check (allow acting on specific statuses even if not the explicit current_approver)
            elif is_hr and obj.status == 'Manager Approved':
                authorized = True
            elif is_finance_exec and obj.status in ['PENDING_EXECUTIVE', 'REJECTED_BY_HEAD', 'PENDING_FINAL_RELEASE', 'HR Approved', 'Approved', 'Under Process']:
                authorized = True
            elif is_finance_head and obj.status in ['PENDING_HEAD', 'PENDING_EXECUTIVE', 'REJECTED_BY_HEAD', 'PENDING_FINAL_RELEASE', 'HR Approved', 'Approved', 'Under Process']:
                authorized = True
            # 3. Finance-specific actions check
            elif is_finance and action in ['Transfer', 'Pay', 'UnderProcess', 'RejectByFinance']:
                authorized = True

            if not authorized:
                raise Exception("You are not authorized to perform this action on this request.")

        from core.models import AuditLog # type: ignore
        if action == 'Reject':
            obj.status = 'Rejected'
            obj.current_approver = None
            obj.save()
            
            AuditLog.objects.create(
                user=user, action='REJECT', model_name=obj.__class__.__name__,
                object_id=str(obj.pk), object_repr=str(obj),
                details={'reason': data.get('remarks') if data else ''}
            )

        if action == 'Forward':
            # This is now mostly handled by 'Approve' automatic progression, 
            # but we keep it for manual overrides if needed by admins.
            mgr_l2 = getattr(obj, 'senior_manager', None)
            mgr_l3 = getattr(obj, 'hod_director', None)
            
            next_approver = None
            next_level = obj.hierarchy_level
            
            if obj.hierarchy_level == 1:
                next_approver = mgr_l2 or mgr_l3
                next_level = 2 if mgr_l2 else 3
            elif obj.hierarchy_level == 2:
                next_approver = mgr_l3
                next_level = 3
            
            if not next_approver:
                raise Exception("There is no higher-level manager to forward this request to.")

            obj.current_approver = next_approver
            obj.hierarchy_level = next_level
            obj.save()
            
            AuditLog.objects.create(
                user=user, action='FORWARD', model_name=obj.__class__.__name__,
                object_id=str(obj.pk), object_repr=str(obj),
                details={'to': str(next_approver)}
            )
            return Response({"message": "Forwarded successfully"})

        if action == 'Approve':
            is_hr = _is_hr(user)
            is_finance = _is_finance_executive(user) or _is_finance_head(user)
            from core.models import AuditLog # type: ignore
            AuditLog.objects.create(
                user=user, action='APPROVE', model_name=obj.__class__.__name__,
                object_id=str(obj.pk), object_repr=str(obj)
            )
            
            # --- STAGE 1: Management Hierarchy ---
            if not is_hr and not is_finance:
                trip = obj if isinstance(obj, Trip) else getattr(obj, 'trip', None)
                if trip:
                    level = getattr(obj, 'hierarchy_level', 1)
                    
                    # If this is a claim, calculate the approved total based on item statuses
                    if isinstance(obj, TravelClaim):
                        from django.db.models import Sum # type: ignore
                        # Re-calculate approved total from items not marked as Rejected
                        # We assume anything not 'Rejected' is 'Approved' or 'Pending' (which defaults to OK in this step)
                        # Actually, better to strictly count only 'Approved' or 'Pending'
                        approved_sum = obj.trip.expenses.exclude(status='Rejected').aggregate(s=Sum('amount'))['s'] or 0
                        obj.approved_amount = approved_sum
                        obj.save()
                        
                        rejected_items = obj.trip.expenses.filter(status='Rejected')
                        rej_msg = f" with {rejected_items.count()} items rejected" if rejected_items.exists() else ""
                        update_trip_lifecycle(trip, f"Level {level} Approval", f"Approved by {user.name}{rej_msg}. Net Approved: ₹{approved_sum}.")
                    else:
                        update_trip_lifecycle(trip, f"Level {level} Approval", f"Approved by {user.name}.")
                
                next_approver = None
                
                # Try explicit levels first
                if obj.hierarchy_level == 1:
                    next_approver = getattr(obj, 'senior_manager', None) or getattr(obj, 'hod_director', None)
                elif obj.hierarchy_level == 2:
                    next_approver = getattr(obj, 'hod_director', None)
                
                # DYNAMIC FALLBACK: If no explicit level but current user has a manager
                if not next_approver:
                    # Use dynamic property to find manager
                    potential_manager = user.reporting_manager
                    
                    # Ensure we don't route back to requester or to a non-existent manager
                    if potential_manager and potential_manager != user and potential_manager != requester:
                        next_approver = potential_manager
                
                if next_approver:
                    # Move to next manager in chain
                    obj.current_approver = next_approver
                    obj.hierarchy_level += 1
                    obj.save()
                    
                    Notification.objects.create(
                        user=next_approver,
                        title=f"Pending Approval: {request_type}",
                        message=f"{requester.name}'s {request_type} requires your review (Forwarded from {user.name}).",
                        type='info'
                    )

                    Notification.objects.create(
                        user=requester,
                        title=f"Approved by {user.name}",
                        message=f"Your {request_type} has been approved by {user.name} and forwarded to {next_approver.name} for the next level review.",
                        type='success'
                    )
                else:
                    # End of Management Chain -> Move to HR Approval
                    hr_head = get_hr_head(requester)
                    obj.status = 'Manager Approved'
                    obj.current_approver = hr_head
                    obj.save()
                    
                    if trip:
                        update_trip_lifecycle(trip, f"Approved by {user.name}", f"Approved by {user.name} and forwarded to HR ({hr_head.name if hr_head else 'Head'}).")
                    
                    Notification.objects.create(
                        user=requester,
                        title=f"Approved by {user.name}",
                        message=f"Your {request_type} has been approved by {user.name} and sent to HR for verification.",
                        type='success'
                    )

                    # Notify Guest House Managers if room request exists and this is a Trip
                    if isinstance(obj, Trip) and obj.accommodation_requests and 'Request for Room' in obj.accommodation_requests:
                        gh_managers = User.objects.filter(role__name='GuestHouseManager')
                        for ghm in gh_managers:
                            Notification.objects.create(
                                user=ghm,
                                title="Pending Room Request",
                                message=f"{requester.name}'s trip to {obj.destination} is management-approved. Room booking may be initiated.",
                                type='info'
                            )
                    
                    if hr_head:
                        Notification.objects.create(
                            user=hr_head,
                            title=f"HR Verification Required",
                            message=f"{requester.name}'s {request_type} has been approved by {user.name} and awaits your verification.",
                            type='info'
                        )
                return Response({"message": "Sent to HR for verification"})

            # --- STAGE 2: HR Approval ---
            elif is_hr:
                trip = obj if isinstance(obj, Trip) else getattr(obj, 'trip', None)
                if trip:
                    update_trip_lifecycle(trip, "Ticket Booking", f"HR ({user.name}) has verified travel logistics.")
                
                if isinstance(obj, Trip):
                    # --- TRIP ENDS AT HR ---
                    obj.status = 'Approved'
                    obj.current_approver = None
                    obj.save()
                    
                    Notification.objects.create(
                        user=requester,
                        title="Trip Approved",
                        message=f"Your Trip to {obj.destination} has been final-approved by HR.",
                        type='success'
                    )

                    # Notify Guest House Managers if room request exists
                    if obj.accommodation_requests and 'Request for Room' in obj.accommodation_requests:
                        gh_managers = User.objects.filter(role__name='GuestHouseManager')
                        for ghm in gh_managers:
                            Notification.objects.create(
                                user=ghm,
                                title="Room Request Ready",
                                message=f"Trip {obj.trip_id} to {obj.destination} has been approved. Room booking is now required.",
                                type='info'
                            )
                else:
                    # --- MONEY REQUESTS MOVE TO FINANCE EXECUTIVE ---
                    # Capture HR specific fields
                    hr_amt = data.get('hr_approved_amount')
                    hr_rem = data.get('hr_remarks') or data.get('remarks')
                    
                    if hasattr(obj, 'hr_approved_amount'):
                        obj.hr_approved_amount = hr_amt if hr_amt is not None else getattr(obj, 'requested_amount', getattr(obj, 'total_amount', 0))
                    if hasattr(obj, 'hr_remarks'):
                        obj.hr_remarks = hr_rem
                    
                    finance_exec = get_finance_executive(requester, exclude_user=user)
                    obj.status = 'PENDING_EXECUTIVE'
                    obj.current_approver = finance_exec
                    obj.save()
                    
                    Notification.objects.create(
                        user=requester,
                        title=f"HR Verified",
                        message=f"Your {request_type} has been verified by HR and sent to Finance Executive for verification.",
                        type='success'
                    )
                    
                    if finance_exec:
                        Notification.objects.create(
                            user=finance_exec,
                            title=f"Finance Verification Required",
                            message=f"{requester.name}'s {request_type} is HR-verified and awaits your verification.",
                            type='info'
                        )
                return Response({"message": "HR recommendation processed"})
            # --- STAGE 3: Finance Approval ---
            elif is_finance:
                trip = obj if isinstance(obj, Trip) else getattr(obj, 'trip', None)
                
                # Case A: Finance Executive Verification (Finance Executive 1)
                if _is_finance_executive(user) and obj.status in ['PENDING_EXECUTIVE', 'HR Approved', 'REJECTED_BY_HEAD', 'Approved', 'Under Process']:
                    exec_amount = data.get('approved_amount') or data.get('executive_approved_amount')
                    if exec_amount is None:
                        exec_amount = getattr(obj, 'requested_amount', getattr(obj, 'total_amount', 0))
                    
                    obj.executive_approved_amount = exec_amount
                    obj.sent_by_executive = user
                    obj.status = 'PENDING_HEAD'
                    obj.current_approver = get_finance_head(requester, exclude_user=user)
                    obj.save()
                    
                    if trip:
                        update_trip_lifecycle(trip, "Finance Verified", f"Verified by Executive {user.name}. Amount recommended: ₹{exec_amount}.")
                    
                    Notification.objects.create(
                        user=obj.current_approver,
                        title="Finance Authorization Required",
                        message=f"{requester.name}'s request verified by executive and awaits your authorization.",
                        type='info'
                    )
                    return Response({"message": "Verified and sent to Head"})

                # Case B: Finance Head Authorization (from PENDING_HEAD)
                if _is_finance_head(user) and obj.status == 'PENDING_HEAD':
                    if action == 'Approve':
                        obj.status = 'PENDING_FINAL_RELEASE'
                        obj.head_action = 'Approved'
                        
                        # Forward to OTHER finance executive (Finance Executive 2)
                        all_fin_users = _get_finance_users()
                        other_execs = [u for u in all_fin_users if not _is_finance_head(u) and u != obj.sent_by_executive]
                        final_exec = other_execs[0] if other_execs else obj.sent_by_executive
                        
                        obj.current_approver = final_exec
                        obj.final_executive = final_exec
                        obj.save()
                        
                        if trip:
                            update_trip_lifecycle(trip, "Finance Authorized", f"Authorized by Head {user.name}. Sent to {final_exec.name if final_exec else 'Executive'} for payout.")
                    
                    elif action == 'Reject':
                        obj.status = 'REJECTED_BY_HEAD'
                        obj.head_action = 'Rejected'
                        obj.current_approver = obj.sent_by_executive
                        obj.save()
                        
                        if trip:
                            update_trip_lifecycle(trip, "Head Rejected", f"Rejected by Finance Head {user.name}. Returned to Executive for revision.")
                    return Response({"message": "Head action processed"})

                # Case C: Transfer action handled below in Finance Specific Actions section
                pass

                return Response({"message": "No specific finance action matched"}, status=200)

        # Finance Specific Actions
        if action == 'UnderProcess':
            obj.status = 'Under Process'
            obj.save()
            Notification.objects.create(
                user=requester,
                title="Finance: Under Process",
                message=f"Your {request_type} is under process by Finance Team.",
                type='info'
            )
            return Response({"message": "Under process updated"})

        if action in ['Transfer', 'Pay']:
            # Security: Only finance can perform transfers, and usually only for finalized requests
            if not _is_finance_executive(user) and not _is_finance_head(user):
                raise PermissionDenied("Only Finance can process transfers.")
            
            # Additional check: If it has a final_executive assigned, only they should pay
            if hasattr(obj, 'final_executive') and obj.final_executive and obj.final_executive != user and not _is_finance_head(user):
                raise PermissionDenied(f"This payout is assigned to {obj.final_executive.name}.")

            if data:
                payment_mode = data.get('payment_mode')
                transaction_id = data.get('transaction_id')
                amount = getattr(obj, 'executive_approved_amount', getattr(obj, 'approved_amount', 0))
                
                # Capture details
                if hasattr(obj, 'payment_mode'): obj.payment_mode = payment_mode
                if hasattr(obj, 'transaction_id'): obj.transaction_id = transaction_id
                if hasattr(obj, 'payment_date'): obj.payment_date = data.get('payment_date') or timezone.now()
                if hasattr(obj, 'finance_remarks'): obj.finance_remarks = data.get('remarks', '')
                if hasattr(obj, 'processed_by'): obj.processed_by = user
            
            # Set target status based on model type
            if isinstance(obj, TravelClaim):
                obj.status = 'Paid'
            else:
                obj.status = 'COMPLETED'
                
            obj.current_approver = None
            obj.save()
            
            # Record Lifecycle Event
            trip = obj if isinstance(obj, Trip) else getattr(obj, 'trip', None)
            if trip:
                # If this is a claim settlement, also update the trip status to 'Settled'
                if isinstance(obj, TravelClaim):
                    trip.status = 'Settled'
                    trip.save()
                update_trip_lifecycle(trip, "Settlement", f"Final Transfer completed by {user.name}.")
            
            Notification.objects.create(
                user=requester,
                title="Amount Credited",
                message=f"Your {request_type} has been fully approved and the amount has been credited to your account.",
                type='success'
            )
            return Response({"message": f"{action} completed and phase closed."})

        if action == 'RejectByFinance':
            obj.status = 'Rejected by Finance'
            reason = data.get('remarks', 'No reason provided') if data else ""
            if hasattr(obj, 'finance_remarks'): obj.finance_remarks = reason
            obj.save()
            Notification.objects.create(
                user=requester,
                title="Finance: Request Rejected",
                message=f"Your {request_type} was rejected by Finance. Reason: {reason}",
                type='error'
            )
            return Response({"message": "Rejected by Finance"})

class ExpenseViewSet(viewsets.ModelViewSet):
    queryset = Expense.objects.all()
    serializer_class = ExpenseSerializer
    permission_classes = [IsCustomAuthenticated]
    http_method_names = ['get', 'post', 'patch', 'put', 'head', 'options']

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return Expense.objects.none()
            
        role_name = user.role.name.lower() if hasattr(user, 'role') else ''
        is_admin = role_name in ['admin', 'superuser']
        is_finance = 'finance' in role_name or role_name == 'cfo'
        is_manager = role_name == 'reporting_authority'
        
        if is_admin or is_finance:
            queryset = self.queryset
        elif is_manager:
            # Allow managers to see their own expenses OR expenses for trips they are/were responsible for
            # We use snapshot names for efficient filtering as dynamic hierarchy lookup is too slow for list views
            queryset = self.queryset.filter(
                Q(trip__user=user) | 
                Q(trip__current_approver=user) |
                Q(trip__reporting_manager_name=user.name) |
                Q(trip__senior_manager_name=user.name) |
                Q(trip__hod_director_name=user.name)
            )
        else:
            queryset = self.queryset.filter(trip__user=user)
            
        trip_id = self.request.query_params.get('trip_id')
        if trip_id:
            trip_id = decode_id(trip_id)
            queryset = queryset.filter(trip__trip_id=trip_id)
        return queryset

    def perform_create(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        trip = serializer.validated_data.get('trip')
        if trip and trip.user != user:
            raise serializers.ValidationError("Unauthorized trip association")
        serializer.save()

class PolicyDocumentViewSet(viewsets.ModelViewSet):
    queryset = PolicyDocument.objects.all()
    serializer_class = PolicyDocumentSerializer
    permission_classes = [IsCustomAuthenticated]

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return PolicyDocumentDetailSerializer
        return PolicyDocumentSerializer

    def get_queryset(self):
        # Always return a fresh queryset; returning the class-level queryset
        # directly can reuse a cached empty result across requests.
        return PolicyDocument.objects.all()

    def perform_create(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        serializer.save(uploaded_by=user)

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            from core.permissions import IsAdmin # type: ignore
            return [IsAdmin()]
        return [IsCustomAuthenticated()]

class TravelClaimViewSet(viewsets.ModelViewSet):
    queryset = TravelClaim.objects.all()
    serializer_class = TravelClaimSerializer
    permission_classes = [IsCustomAuthenticated]

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return TravelClaim.objects.none()
            
        is_admin = hasattr(user, 'role') and user.role.name.lower() in ['admin', 'superuser']
        
        if is_admin:
            queryset = self.queryset
        else:
            queryset = self.queryset.filter(trip__user=user)
            
        trip_id = self.request.query_params.get('trip_id')
        if trip_id:
            trip_id = decode_id(trip_id)
            queryset = queryset.filter(trip__trip_id=trip_id)
        return queryset

    def perform_create(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            raise serializers.ValidationError("Authentication failed. Please log in again.")
            
        trip = serializer.validated_data.get('trip')
        if trip and trip.user != user:
            raise serializers.ValidationError("Unauthorized trip association")
        
        reporting_manager = user.reporting_manager
        senior_manager = user.senior_manager
        hod_director = user.hod_director
        
        # Logic to find first available approver
        current_approver = reporting_manager
        h_level = 1
        
        if not current_approver:
            current_approver = senior_manager
            h_level = 2
        
        if not current_approver:
            current_approver = hod_director
            h_level = 3
            
        if not current_approver:
            # If no managers at all, go to HR
            current_approver = get_hr_head(user)

        from django.db.models import Sum # type: ignore
        total_expense_sum = trip.expenses.aggregate(s=Sum('amount'))['s'] or 0

        claim = serializer.save(
            status='Submitted',
            current_approver=current_approver,
            hierarchy_level=h_level,
            submitted_at=timezone.now(),
            total_amount=total_expense_sum,
            approved_amount=total_expense_sum,
            # Populate snapshots for resilience
            user_name=user.name,
            user_designation=user.designation,
            user_department=user.department,
            reporting_manager_name=reporting_manager.name if reporting_manager else None,
            senior_manager_name=senior_manager.name if senior_manager else None,
            hod_director_name=hod_director.name if hod_director else None
        )
        
        if current_approver:
            Notification.objects.create(
                user=current_approver,
                title="New Expense Claim",
                message=f"{user.name} has submitted an expense claim for Trip {claim.trip.trip_id}.",
                type='info'
            )
            
        # Notify HR
        notify_hr("New Expense Claim", f"{user.name} has submitted an expense claim for Trip {claim.trip.trip_id}.")
            
        self._update_trip_lifecycle(claim.trip)

    def perform_update(self, serializer):
        claim = serializer.save()
        
        from django.db.models import Sum # type: ignore
        total_amount = claim.trip.expenses.aggregate(s=Sum('amount'))['s'] or 0
        claim.total_amount = total_amount
        claim.approved_amount = total_amount
        claim.save()
        
        if claim.status == 'Submitted':
            self._update_trip_lifecycle(claim.trip)

    def _update_trip_lifecycle(self, trip):
        update_trip_lifecycle(trip, "Settlement", "Travel reimbursement claim submitted for review.")

class TravelAdvanceViewSet(viewsets.ModelViewSet):
    queryset = TravelAdvance.objects.all()
    serializer_class = TravelAdvanceSerializer

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return TravelAdvance.objects.none()
            
        is_admin = hasattr(user, 'role') and user.role.name.lower() in ['admin', 'superuser']
        
        if is_admin:
            queryset = self.queryset
        else:
            queryset = self.queryset.filter(trip__user=user)
            
        trip_id = self.request.query_params.get('trip_id')
        if trip_id:
            trip_id = decode_id(trip_id)
            queryset = queryset.filter(trip__trip_id=trip_id)
        return queryset
    def perform_create(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        if not user:
             from rest_framework.exceptions import AuthenticationFailed # type: ignore
             raise AuthenticationFailed("Action requires authentication.")

        trip = serializer.validated_data.get('trip')
        if trip and trip.user != user:
            raise serializers.ValidationError("Unauthorized trip association")

        reporting_manager = user.reporting_manager
        senior_manager = user.senior_manager
        hod_director = user.hod_director

        # Logic to find first available approver
        current_approver = reporting_manager
        h_level = 1
        
        if not current_approver:
            current_approver = senior_manager
            h_level = 2
        
        if not current_approver:
            current_approver = hod_director
            h_level = 3
            
        if not current_approver:
            # If no managers at all, go to HR
            current_approver = get_hr_head(user)

        advance = serializer.save(
            status='Submitted',
            current_approver=current_approver,
            hierarchy_level=h_level,
            submitted_at=timezone.now(),
            # Populate snapshots for resilience
            user_name=user.name,
            user_designation=user.designation,
            user_department=user.department,
            reporting_manager_name=reporting_manager.name if reporting_manager else None,
            senior_manager_name=senior_manager.name if senior_manager else None,
            hod_director_name=hod_director.name if hod_director else None
        )
        
        if current_approver:
            Notification.objects.create(
                user=current_approver,
                title="New Advance Request",
                message=f"{user.name} has requested an advance of ₹{advance.requested_amount} for Trip {advance.trip.trip_id}.",
                type='info'
            )
            
        # Notify HR
        notify_hr("New Advance Request", f"{user.name} has requested an advance of ₹{advance.requested_amount} for Trip {advance.trip.trip_id}.")

        self._update_trip_lifecycle(advance)

    def perform_update(self, serializer):
        advance = serializer.save()
        if advance.status == 'Submitted':
            self._update_trip_lifecycle(advance)

    def _update_trip_lifecycle(self, advance):
        update_trip_lifecycle(advance.trip, "Advance Requested", f"Pre-trip advance of ₹{advance.requested_amount} requested for: {advance.purpose[:40]}...")



class DisputeViewSet(viewsets.ModelViewSet):
    queryset = Dispute.objects.all()
    serializer_class = DisputeSerializer
    permission_classes = [IsCustomAuthenticated]

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return Dispute.objects.none()
        
        if hasattr(user, 'role') and user.role.name.lower() in ['admin', 'superuser']:
            return Dispute.objects.all()
        
        return Dispute.objects.filter(raised_by=user)

    def perform_create(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        serializer.save(raised_by=user)

class TripOdometerViewSet(viewsets.ModelViewSet):
    queryset = TripOdometer.objects.all()
    serializer_class = TripOdometerSerializer
    permission_classes = [IsCustomAuthenticated]

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return TripOdometer.objects.none()
            
        is_admin = hasattr(user, 'role') and user.role.name.lower() in ['admin', 'superuser']
        
        if is_admin:
            queryset = self.queryset
        else:
            queryset = self.queryset.filter(trip__user=user)
            
        trip_id = self.request.query_params.get('trip_id')
        if trip_id:
            trip_id = decode_id(trip_id)
            queryset = queryset.filter(trip__trip_id=trip_id)
        return queryset

    def perform_create(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        trip = serializer.validated_data.get('trip')
        if trip and trip.user != user:
            raise serializers.ValidationError("Unauthorized trip association")
            
            
        odo = serializer.save()
        if odo.start_odo_reading:
            trip = odo.trip
            trip.status = 'On-Going'
            trip.save(update_fields=['status'])
            update_trip_lifecycle(trip, "Journey Started", f"Trip journey started with odometer reading {odo.start_odo_reading} KM.")

    def perform_update(self, serializer):
        odo = serializer.save()
        if odo.end_odo_reading:
            trip = odo.trip
            trip.status = 'Completed'
            trip.save(update_fields=['status'])
            update_trip_lifecycle(trip, "Journey Ended", f"Trip journey completed with final odometer reading {odo.end_odo_reading} KM.")

class DashboardStatsView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        user = getattr(request, 'custom_user', None)
        if not user:
            return Response({"error": "User not found"}, status=401)
        
        is_admin = user.role.name.lower() == 'admin'
        is_gh_manager = user.role.name.lower() == 'guesthousemanager'
        is_fin_head = _is_finance_head(user)
        is_fin_exec = _is_finance_executive(user)
        is_finance = is_fin_head or is_fin_exec
        is_cfo = 'cfo' in (user.role.name.lower() if user.role else '')
        is_hr = _is_hr(user)

        if is_admin or is_gh_manager or is_finance or is_cfo:
            trips = Trip.objects.all()
            base_expenses = Expense.objects.all()
        else:
            trips = Trip.objects.filter(user=user)
            base_expenses = Expense.objects.filter(trip__user=user)

        total_trips = trips.count()
        in_review = trips.filter(status='Pending').count()
        
        # Count tasks awaiting this user's approval
        pending_action = 0
        # variables already defined above

        if not is_admin:
            pending_action += Trip.objects.filter(current_approver=user).count()
            pending_action += TravelAdvance.objects.filter(current_approver=user).count()
            pending_action += TravelClaim.objects.filter(current_approver=user).count()
            
            # For Finance, count specific stages they act on
            if is_fin_head:
                pending_action += TravelAdvance.objects.filter(status='PENDING_HEAD').count()
                pending_action += TravelClaim.objects.filter(status='PENDING_HEAD').count()
            elif is_fin_exec:
                money_statuses = ['PENDING_EXECUTIVE', 'REJECTED_BY_HEAD', 'PENDING_FINAL_RELEASE']
                pending_action += TravelAdvance.objects.filter(status__in=money_statuses).count()
                pending_action += TravelClaim.objects.filter(status__in=money_statuses).count()
            elif is_hr:
                pending_action += Trip.objects.filter(status='Manager Approved').count()
            elif is_gh_manager:
                # For GH Manager, count trips that need room booking
                pending_action += Trip.objects.filter(
                    accommodation_requests__contains='Request for Room',
                    status__in=['Manager Approved', 'Approved']
                ).exclude(room_bookings__isnull=False).distinct().count()
        
        total_expenses = base_expenses.aggregate(Sum('amount'))['amount__sum'] or 0
        
        # compute total approved advances for the selected trips - used for wallet/advance balances
        # mirror TripSerializer.get_total_approved_advance logic (consider executive_approved_amount when >0)
        total_approved_advances = 0.0
        advances_qs = TravelAdvance.objects.filter(
            trip__in=trips,
            status__in=['Paid', 'Transferred', 'COMPLETED']
        )
        for adv in advances_qs:
            amt = float(adv.executive_approved_amount) if float(adv.executive_approved_amount) > 0 else float(adv.requested_amount)
            total_approved_advances += amt
        
        wallet_balance = float(total_approved_advances) - float(total_expenses)
        
        approved_expenses_qs = base_expenses.filter(
            Q(trip__claim__status__in=['Approved', 'Paid']) |
            Q(trip__status__in=['Approved', 'Completed', 'Settled'], trip__consider_as_local=True)
        )
        approved_expenses = approved_expenses_qs.aggregate(Sum('amount'))['amount__sum'] or 0
        
        categories = base_expenses.values('category').annotate(total=Sum('amount'))
        
        recent_trips = trips.order_by('-created_at')[:5]
        recent_data = []
        for t in recent_trips:
            # If trip is approved/completed, show actual total instead of estimate
            if t.status in ['Approved', 'Completed', 'Settled']:
                actual_total = Expense.objects.filter(trip=t).aggregate(Sum('amount'))['amount__sum'] or 0
                display_amount = f"₹{actual_total:,.0f}" if actual_total > 0 else t.cost_estimate
            else:
                display_amount = t.cost_estimate

            recent_data.append({
                "id": t.trip_id,
                "title": f"{'Travel' if t.consider_as_local else 'Trip'} to {t.destination}",
                "subtitle": f"{t.user.name} - {t.purpose}" if (is_admin or is_gh_manager) and t.user else t.purpose,
                "status": t.status,
                "amount": display_amount
            })

        kpis = [
            { "title": 'Total Trips', "value": str(total_trips), "label": 'Managed trips' if is_admin or is_finance or is_gh_manager else 'Your trips', "icon": 'Briefcase', "color": 'orange' },
            { "title": 'Approved Expenses', "value": f"₹{approved_expenses:,.0f}", "label": 'Finalized' if is_finance else 'Confirmed', "icon": 'CreditCard', "color": 'red' },
            { "title": 'Total Spend' if not is_finance else 'Total Disbursements', "value": f"₹{total_expenses:,.0f}", "label": 'Recorded', "icon": 'TrendingUp', "color": 'magenta' },
            { 
                "title": 'Action Required' if pending_action > 0 else 'In Review', 
                "value": str(pending_action if pending_action > 0 else in_review), 
                "label": 'Pending your action' if pending_action > 0 else 'Pending trips', 
                "icon": 'Clock', 
                "color": 'yellow' 
            }
        ]

        if is_finance:
            # Shift to vibrant functional colors (strictly within supported index.css classes)
            kpis[0]['color'] = 'orange'
            kpis[1]['color'] = 'magenta'
            kpis[1]['title'] = 'Net Payouts'
            kpis[2]['color'] = 'red'

        return Response({
            "kpis": kpis,
            "recent_activity": recent_data,
            "is_finance_hub": is_finance,
            "expenditure_mix": [
                { 
                    "type": dict(Expense.CATEGORY_CHOICES).get(cat['category'], cat['category']), 
                    "amount": float(cat['total']), 
                    "percentage": (float(cat['total']) / float(total_expenses) * 100) if total_expenses > 0 else 0 
                }
                for cat in categories
            ],
            "total_spend": total_expenses,
            # wallet_balance represents approved advances minus expenses across the user/trips
            "wallet_balance": wallet_balance,
            # advance_balance is simply the sum of approved advances
            "advance_balance": float(total_approved_advances)
        })

class TripSettlementView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        user = getattr(request, 'custom_user', None)
        trip_id_enc = request.query_params.get('trip_id')
        
        if not trip_id_enc:
            # If no trip_id provided, return all trips that are ready for settlement
            # (i.e. trips with approved claims that aren't PAID/COMPLETED yet)
            # For simplicity, returning a list of all trips the user can see for now
            # but we filter for finance if they are finance
            is_finance = _is_finance_executive(user) or _is_finance_head(user)
            if is_finance:
                trips = Trip.objects.filter(status__in=['Claim Submitted', 'Manager Approved', 'Approved'])
            else:
                trips = Trip.objects.filter(user=user)
            
            data = []
            for t in trips:
                advances = TravelAdvance.objects.filter(trip=t, status='COMPLETED').aggregate(Sum('requested_amount'))['requested_amount__sum'] or 0
                claim_amt = 0
                if hasattr(t, 'claim'):
                    claim_amt = t.claim.total_amount
                
                data.append({
                    "trip_id": t.trip_id,
                    "destination": t.destination,
                    "employee": t.user.name if t.user else "Unknown",
                    "advance": float(advances),
                    "claim": float(claim_amt),
                    "balance": float(claim_amt - advances),
                    "status": t.status
                })
            return Response(data)

        trip_id = decode_id(trip_id_enc)
        try:
            trip = Trip.objects.get(trip_id=trip_id)
        except Trip.DoesNotExist:
            return Response({"error": "Trip not found"}, status=404)

        advances_list = TravelAdvance.objects.filter(trip=trip).order_by('created_at')
        total_advance = sum(a.requested_amount for a in advances_list if a.status == 'COMPLETED')
        
        claim_amt = 0
        claim_id = None
        claim_date = None
        if hasattr(trip, 'claim'):
            claim_amt = trip.claim.total_amount
            claim_id = f"CLAIM-{trip.claim.id}"
            claim_date = trip.claim.submitted_at.strftime("%B %d, %Y") if trip.claim.submitted_at else None

        breakdown = []
        for adv in advances_list:
            if adv.status == 'COMPLETED':
                breakdown.append({
                    "id": f"ADV-{adv.id}",
                    "type": "Advance",
                    "description": f"Advance via {adv.payment_mode or 'Bank Transfer'}",
                    "date": adv.payment_date.strftime("%B %d, %Y") if adv.payment_date else adv.created_at.strftime("%B %d, %Y"),
                    "amount": -float(adv.requested_amount),
                    "is_negative": True
                })

        if claim_amt > 0:
            breakdown.append({
                "id": claim_id,
                "type": "Claim",
                "description": f"Expense Claim: {trip.trip_id}",
                "date": claim_date or trip.updated_at.strftime("%B %d, %Y"),
                "amount": float(claim_amt),
                "is_negative": False
            })

        balance = float(claim_amt - total_advance)
        
        return Response({
            "summary": {
                "advance": float(total_advance),
                "claimTotal": float(claim_amt),
                "balance": balance,
                "type": "Payable" if balance > 0 else "Recoverable",
                "status": trip.status
            },
            "breakdown": breakdown,
            "trip": {
                "id": trip.trip_id,
                "destination": trip.destination,
                "employee": trip.user.name if trip.user else "Unknown"
            }
        })

    def post(self, request):
        # Handle Finalize & Settle action
        user = getattr(request, 'custom_user', None)
        if not user:
            return Response({"error": "Authentication required"}, status=401)
            
        trip_id_enc = request.data.get('trip_id')
        if not trip_id_enc:
            return Response({"error": "Trip ID required"}, status=400)
            
        trip_id = decode_id(trip_id_enc)
        try:
            trip = Trip.objects.get(trip_id=trip_id)
        except Trip.DoesNotExist:
            return Response({"error": "Trip not found"}, status=404)

        # Logic to finalize
        trip.status = 'Settled'
        trip.save()
        
        if hasattr(trip, 'claim'):
            trip.claim.status = 'Paid'
            trip.claim.save()

        update_trip_lifecycle(trip, "Final Settlement", f"Accounts settled and closed by {user.name}.")
        
        return Response({"message": "Trip account successfully settled."})

class CFOWarRoomView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        now = timezone.now()
        first_day_of_current_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # 1. Monthly Total Spend
        monthly_spend = Expense.objects.filter(date__gte=first_day_of_current_month.date()).aggregate(Sum('amount'))['amount__sum'] or 0
        
        # Comparison with last month
        last_month_end = first_day_of_current_month - timezone.timedelta(seconds=1)
        last_month_start = last_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_month_spend = Expense.objects.filter(date__gte=last_month_start.date(), date__lte=last_month_end.date()).aggregate(Sum('amount'))['amount__sum'] or 0
        
        percent_change_spend = 0
        if last_month_spend > 0:
            percent_change_spend = ((float(monthly_spend) - float(last_month_spend)) / float(last_month_spend)) * 100
        
        # 2. Avg Cost Per Trip
        total_claims = TravelClaim.objects.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        total_trips = Trip.objects.count()
        avg_cost = (float(total_claims) / total_trips) if total_trips > 0 else 0
        
        # 3. Guest House Occupancy
        from guest_house.models import Room, RoomBooking # type: ignore
        total_rooms = Room.objects.count()
        # Active bookings today
        active_bookings = RoomBooking.objects.filter(start_date__lte=now, end_date__gte=now).count()
        occupancy = (active_bookings / total_rooms * 100) if total_rooms > 0 else 0
        
        # 4. Policy Overrides/Disputes
        # Using Category 'Policy' in Dispute as a proxy for overrides
        policy_issues = Dispute.objects.filter(category='Policy').count()
        policy_rate = (policy_issues / total_trips * 100) if total_trips > 0 else 0

        # KPI Stats Construction
        stats = [
            { 
                "title": 'Total Spend (Monthly)', 
                "value": f"₹{monthly_spend:,.0f}", 
                "change": f"{abs(float(percent_change_spend)):.1f}%", # type: ignore
                "trend": 'up' if percent_change_spend >= 0 else 'down',
                "icon": 'IndianRupee'
            },
            { 
                "title": 'Avg Cost per Trip', 
                "value": f"₹{avg_cost:,.0f}", 
                "change": '+2.4%', # Placeholder for complexity
                "trend": 'up', 
                "icon": 'TrendingUp' 
            },
            { 
                "title": 'Guest House Occupancy', 
                "value": f"{occupancy:.1f}%", 
                "change": '+1.8%', 
                "trend": 'up', 
                "icon": 'Building2' 
            },
            { 
                "title": 'Policy Overrides', 
                "value": f"{policy_rate:.1f}%", 
                "change": '-0.2%', 
                "trend": 'down', 
                "icon": 'AlertCircle' 
            },
        ]

        # 5. Spend by Department
        dept_raw = Expense.objects.values('trip__user_department').annotate(total=Sum('amount')).order_by('-total')
        spend_by_dept = [
            { "name": d['trip__user_department'] or "Unknown", "value": float(d['total']) }
            for d in dept_raw
        ]

        # 6. Advance Aging
        advances = TravelAdvance.objects.exclude(status='COMPLETED')
        aging_buckets = { "0-30": 0.0, "31-60": 0.0, "60+": 0.0 }
        for adv in advances:
            days = (now - adv.created_at).days
            amt = float(adv.requested_amount)
            if days <= 30: aging_buckets["0-30"] += amt
            elif days <= 60: aging_buckets["31-60"] += amt
            else: aging_buckets["60+"] += amt
            
        aging_data = [
            { "range": '0-30 Days', "amount": aging_buckets["0-30"], "color": 'success' },
            { "range": '31-60 Days', "amount": aging_buckets["31-60"], "color": 'warning' },
            { "range": '60+ Days', "amount": aging_buckets["60+"], "color": 'danger' },
        ]

        # 7. Critical Anomalies
        anomalies = [
            { "entity": "Logistics Surge", "reason": "Local Conveyance in Metro areas up by 40%", "impact": "High", "action": "Review" },
            { "entity": "Dept Override", "reason": "Bulk travel booking policy bypassed in IT dept", "impact": "Medium", "action": "Investigate" }
        ]

        return Response({
            "stats": stats,
            "report_month": now.strftime("%B %Y"),
            "spend_by_dept": spend_by_dept,
            "aging": aging_data,
            "anomalies": anomalies
        })

import json

class BulkActivityBatchViewSet(viewsets.ModelViewSet):
    queryset = BulkActivityBatch.objects.all()
    serializer_class = BulkActivityBatchSerializer
    permission_classes = [IsCustomAuthenticated]

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user: return BulkActivityBatch.objects.none()
        
        role_name = (user.role.name if user.role else '').lower()
        
        # Admins/Finance/COO see all
        if any(kw in role_name for kw in ['admin', 'finance', 'cfo', 'coo']):
            return self.queryset
        
        # Allow any user to see batches they submitted OR batches they need to approve.
        # This ensures that managers (like a COO) can see and approve their team's uploads 
        # even if their system role is simply 'Employee'.
        return self.queryset.filter(Q(user=user) | Q(current_approver=user)).order_by('-created_at')

    @action(detail=False, methods=['get'])
    def template(self, request):
        import openpyxl # type: ignore
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side # type: ignore
        from openpyxl.utils import get_column_letter # type: ignore
        from openpyxl.worksheet.datavalidation import DataValidation # type: ignore
        from openpyxl.styles.differential import DifferentialStyle # type: ignore
        from openpyxl.formatting.rule import FormulaRule # type: ignore
        import datetime
        from django.utils import timezone # type: ignore

        wb = openpyxl.Workbook()

        # ── Hidden locations sheet ───────────────────────────────────────────
        loc_sheet = wb.create_sheet("_Locations")
        loc_sheet.sheet_state = 'hidden'

        # 1. Try to get locations of employees reporting to the current manager
        from api_management.services import get_manager_reports_locations # type: ignore
        user = getattr(request, 'custom_user', None)
        manager_code = user.employee_id if user else None
        
        # Gather clusters from all employees in the reporting chain
        team_locs = set()
        if manager_code:
            for loc in get_manager_reports_locations(manager_code):
                if loc:
                    team_locs.add(loc)

        # Add the manager's own cluster/district (using geo_location priority)
        if user:
            own_loc = (user.office_location or '').strip()
            if own_loc:
                team_locs.add(own_loc)

        # Build final sorted list – strictly team clusters only, no trip history
        if team_locs:
            locations = sorted(team_locs)
        else:
            locations = ["Head Office", "Field Office", "Client Site"]

        for i, loc in enumerate(locations, start=1):
            loc_sheet.cell(row=i, column=1, value=loc)

        # Create a named range for the locations
        loc_range = f"_Locations!$A$1:$A${len(locations)}"
        wb.defined_names['LocationList'] = openpyxl.workbook.defined_name.DefinedName(
            'LocationList', attr_text=loc_range
        )

        # ── Main data sheet ─────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Monthly Activities"

        # ── Styles ──────────────────────────────────────────────────────────
        HEADER_FILL   = PatternFill("solid", fgColor="BB0633")   # brand red
        NOTE_FILL     = PatternFill("solid", fgColor="FFF3CD")   # warm yellow
        HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        NOTE_FONT   = Font(name="Calibri", italic=True, color="856404", size=9)
        DATA_FONT     = Font(name="Calibri", size=10)
        CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin          = Side(style="thin", color="CCCCCC")
        BORDER        = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Column layout ────────────────────────────────────────────────────
        # A=Date  B=From Location  C=Start Time  D=To Location  E=Reach Time  F=Purpose
        columns = [
            ("Date",          16, CENTER),
            ("From Location", 28, LEFT),
            ("Start Time",    14, CENTER),
            ("To Location",   28, LEFT),
            ("Reach Time",    14, CENTER),
            ("Purpose",       40, LEFT),
        ]

        # Row 1 – Headers
        ws.row_dimensions[1].height = 30
        for col_idx, (label, width, align) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font   = HEADER_FONT
            cell.fill   = HEADER_FILL
            cell.alignment = CENTER
            cell.border    = BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Row 2 – Instructional note (merged across all columns)
        ws.merge_cells("A2:F2")
        note_cell = ws.cell(row=2, column=1,
            value="📋  Instructions: Date must be ≥ today  |  Start Time & Reach Time must be HH:MM (24h)  |"
                  " Reach Time must be > Start Time | "
                  "Choose From/To Location from the dropdown  |  Purpose is free text")
        note_cell.font      = NOTE_FONT
        note_cell.fill      = NOTE_FILL
        note_cell.alignment = LEFT
        ws.row_dimensions[2].height = 22

        # Row 3 – Sample data row
        # Values for pre-filled row 3
        # Explicitly convert to IST (UTC+5:30) — server TIME_ZONE is UTC
        from zoneinfo import ZoneInfo
        ist_tz    = ZoneInfo('Asia/Kolkata')
        now_ist   = datetime.datetime.now(ist_tz)
        
        # Round time to nearest 5 minutes
        minute_rounded = (now_ist.minute // 5) * 5
        now_rounded = now_ist.replace(minute=minute_rounded, second=0, microsecond=0)

        # Use actual date/time objects so validation formulas can compare them
        today_obj = datetime.date.today()
        # Ensure Row 3 sample data is treated as date/time
        sample_data = [
            today_obj,
            locations[0] if locations else "Office",
            now_rounded.time(), 
            locations[1] if len(locations) > 1 else "Client Site",
            (now_rounded + datetime.timedelta(hours=1)).time(),
            "Site Inspection / Field Visit"
        ]
        ws.row_dimensions[3].height = 20
        for col_idx, val in enumerate(sample_data, start=1):
            cell = ws.cell(row=3, column=col_idx, value=val)
            cell.font      = DATA_FONT
            cell.alignment = columns[col_idx - 1][2]
            cell.border    = BORDER
            # Set explicit formats for A and B
            if col_idx == 1: cell.number_format = 'YYYY-MM-DD'
            if col_idx == 2: cell.number_format = 'HH:MM'

        # Rows 4-103 – Empty data rows (100 rows)
        for row in range(4, 104):
            ws.row_dimensions[row].height = 20
            for col_idx in range(1, 7):
                cell = ws.cell(row=row, column=col_idx)
                cell.font      = DATA_FONT
                cell.alignment = columns[col_idx - 1][2]
                cell.border    = BORDER
                # Column A = Date, Column C = Start Time, Column E = Reach Time
                if col_idx == 1: cell.number_format = 'YYYY-MM-DD'
                if col_idx in [3, 5]: cell.number_format = 'HH:MM'

        DATA_ROWS = "3:103"   # applies to sample + 100 blank rows

        # ── Validation 1: Date >= today ──────────────────────────────────────
        today_serial = (datetime.date.today() - datetime.date(1899, 12, 30)).days
        dv_date = DataValidation(
            type="date",
            operator="greaterThanOrEqual",
            formula1=str(today_serial),
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid Date",
            error=f"Date must be on or after today ({datetime.date.today().strftime('%d-%b-%Y')}). "
                  f"Please enter a valid date.",
            showInputMessage=True,
            promptTitle="Date",
            prompt=f"Enter date ≥ {datetime.date.today().strftime('%d-%b-%Y')} (YYYY-MM-DD format)"
        )
        ws.add_data_validation(dv_date)
        dv_date.sqref = "A3:A103"   # Column A

        # -- Updated formula: 5-min increments AND (if Date=Today, Time>Now; else Date>Today) --
        # Now column C
        start_time_formula = '=AND(MOD(ROUND(C3*1440,0),5)=0, IF(A3=TODAY(), C3>MOD(NOW(),1), A3>TODAY()))'
        
        dv_start_time = DataValidation(
            type="custom",
            formula1=start_time_formula,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid Start Time",
            error="Start Time must be a multiple of 5 minutes. If selected date is today, time must be > current time.",
            showInputMessage=True,
            promptTitle="Start Time (HH:MM)",
            prompt="Enter time in HH:MM (24h). If date is today, it must be > current time. Must be a multiple of 5 minutes."
        )
        ws.add_data_validation(dv_start_time)
        dv_start_time.sqref = "C3:C103"   # Column C

        # -- Reach Time Validation: Multiple of 5 mins AND Reach Time > Start Time --
        # Now column E, compared with C
        reach_time_formula = '=AND(MOD(ROUND(E3*1440,0),5)=0, E3>C3)'
        dv_reach_time = DataValidation(
            type="custom",
            formula1=reach_time_formula,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid Reach Time",
            error="Reach Time must be a multiple of 5 minutes and must be GREATER than Start Time.",
            showInputMessage=True,
            promptTitle="Reach Time (HH:MM)",
            prompt="Enter time in HH:MM (24h). Must be > Start Time and a multiple of 5 minutes."
        )
        ws.add_data_validation(dv_reach_time)
        dv_reach_time.sqref = "E3:E103"   # Column E

        # ── Validation 3: From Location dropdown ─────────────────────────────
        # Now column B
        dv_from = DataValidation(
            type="list",
            formula1=f'_Locations!$A$1:$A${len(locations)}',
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid Location",
            error="Please select a location from the dropdown list.",
            showInputMessage=True,
            promptTitle="From Location",
            prompt="Select the starting location from the dropdown."
        )
        ws.add_data_validation(dv_from)
        dv_from.sqref = "B3:B103"   # Column B

        # ── Validation 4: To Location dropdown ───────────────────────────────
        # Now column D
        dv_to = DataValidation(
            type="list",
            formula1=f'_Locations!$A$1:$A${len(locations)}',
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid Location",
            error="Please select a location from the dropdown list.",
            showInputMessage=True,
            promptTitle="To Location",
            prompt="Select the destination location from the dropdown."
        )
        ws.add_data_validation(dv_to)
        dv_to.sqref = "D3:D103"   # Column D

        # Highlight Red if From == To (B and D)
        from openpyxl.styles import PatternFill # type: ignore
        red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        ws.conditional_formatting.add("B3:D103",
            openpyxl.formatting.rule.FormulaRule(formula=["=$B3=$D3"], stopIfTrue=True, fill=red_fill))


        # Column E (Purpose) – no list validation, free text; just an input hint
        dv_purpose = DataValidation(
            type="textLength",
            operator="greaterThan",
            formula1="0",
            showErrorMessage=True,
            errorTitle="Purpose Required",
            error="Please describe the purpose of your visit.",
            showInputMessage=True,
            promptTitle="Purpose",
            prompt="Briefly describe the reason for this visit (e.g. Site Inspection, Client Meeting)."
        )
        ws.add_data_validation(dv_purpose)
        dv_purpose.sqref = "F3:F103"   # Column F

        # ── Freeze panes below header + note rows ────────────────────────────
        ws.freeze_panes = "A3"

        # ── Output ───────────────────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="bulk_activity_template.xlsx"'
        return response

    def get_queryset(self):
        queryset = BulkActivityBatch.objects.all().order_by('-created_at')
        trip_id = self.request.query_params.get('trip_id')
        if trip_id:
            queryset = queryset.filter(trip_id=trip_id)
        
        user = getattr(self.request, 'custom_user', None)
        if user:
            from django.db.models import Q # type: ignore
            queryset = queryset.filter(Q(user=user) | Q(current_approver=user))
            
        return queryset

    @action(detail=False, methods=['post'])
    def upload(self, request):
        user = getattr(request, 'custom_user', None)
        if not user:
            return Response({"error": "Authentication required"}, status=401)

        file = request.FILES.get('file')
        json_data = request.data.get('data_json') # For resubmitting/editing rejected rows
        trip_id = request.data.get('trip_id') # Selected by user in UI
        parent_batch_id = request.data.get('parent_batch_id')
        
        if not file and not json_data:
            return Response({"error": "No file or data provided"}, status=400)
            
        rows = []
        file_name = "resubmission.json"
        
        try:
            if file:
                file_name = file.name
                # drop the instruction row that sits just below the header
                df = pd.read_excel(file, skiprows=[1])

                for index, row in df.iterrows():
                    if row.dropna().empty: continue
                    date_raw = row.get('Date', row.get('Date (YYYY-MM-DD)', ''))
                    if pd.isna(date_raw): continue
                    date_val = str(date_raw).strip()
                    if 'instruc' in date_val.lower(): continue
                    if len(str(date_val)) > 10: date_val = str(date_val)[:10] # type: ignore

                    start_time_raw = row.get('Start Time', row.get('Time', ''))
                    start_time_str = '09:00'
                    if start_time_raw and str(start_time_raw).strip() not in ('', 'nan', 'NaT'):
                        start_time_str = str(start_time_raw).strip()
                        try:
                            frac = float(start_time_str)
                            if 0 <= frac < 1:
                                total_secs = int(frac * 86400)
                                start_time_str = f"{total_secs // 3600:02d}:{(total_secs % 3600) // 60:02d}"
                        except (ValueError, TypeError): pass
                        start_time_str = ':'.join(str(start_time_str).split(':')[:2]) # type: ignore

                    reach_time_raw = row.get('Reach Time', row.get('End Time', ''))
                    reach_time_str = '18:00'
                    if reach_time_raw and str(reach_time_raw).strip() not in ('', 'nan', 'NaT'):
                        reach_time_str = str(reach_time_raw).strip()
                        try:
                            frac = float(reach_time_str)
                            if 0 <= frac < 1:
                                total_secs = int(frac * 86400)
                                reach_time_str = f"{total_secs // 3600:02d}:{(total_secs % 3600) // 60:02d}"
                        except (ValueError, TypeError): pass
                        reach_time_str = ':'.join(str(reach_time_str).split(':')[:2]) # type: ignore

                    origin = str(row.get('From Location', row.get('from location', ''))).strip()
                    destination = str(row.get('To Location', row.get('to location', ''))).strip()
                    if origin and destination and origin.lower() == destination.lower():
                        raise Exception(f"Row {index + 3}: From Location and To Location cannot be the same.")

                    rows.append({
                        "date": date_val, "start_time": start_time_str, "reach_time": reach_time_str, "mode": "Bike",
                        "origin_route": origin, "destination_route": destination,
                        "vehicle": "Own Bike", "visit_intent": str(row.get('Purpose', row.get('purpose', '')))
                    })
            else:
                # Direct JSON submission (resubmitting rejected rows)
                rows = json_data if isinstance(json_data, list) else []
                from datetime import datetime
                file_name = f"resubmitted_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
            
            if not rows:
                return Response({"error": "No valid activity rows found"}, status=400)

            if not user:
                return Response({"error": "User not found in request"}, status=400)

            rm, sm, hod = user.reporting_manager, user.senior_manager, user.hod_director
            
            def is_mgmt(u):
                return u and hasattr(u, 'role') and u.role and u.role.name.lower() in ['admin', 'superuser', 'it admin']

            current_approver = rm if not is_mgmt(rm) else (sm if not is_mgmt(sm) else (hod if not is_mgmt(hod) else None))
            h_level = 1 if current_approver == rm else (2 if current_approver == sm else (3 if current_approver == hod else 1))
            
            if not current_approver:
                 current_approver = get_hr_head(user)
                 h_level = 1

            batch = BulkActivityBatch.objects.create(
                user=user, trip_id=trip_id, file_name=file_name,
                data_json=rows, status='Submitted',
                current_approver=current_approver, hierarchy_level=h_level
            )
            
            # If this is a resubmission, mark the parent batch as Resolved
            if parent_batch_id:
                try:
                    parent_batch = BulkActivityBatch.objects.get(id=parent_batch_id)
                    parent_batch.status = 'Resolved'
                    parent_batch.save()
                except BulkActivityBatch.DoesNotExist:
                    pass
            
            if current_approver:
                Notification.objects.create(
                    user=current_approver, title="New Bulk Activity Batch",
                    message=f"{user.name} submitted a bulk travel log for approval.", type='info'
                )
            
            return Response(BulkActivityBatchSerializer(batch).data)
        except Exception as e:
            return Response({"error": f"Failed to process submission: {str(e)}"}, status=400)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve_batch(self, request, pk=None):
        batch = self.get_object()
        user = getattr(request, 'custom_user', None)
        
        if not user:
             return Response({"error": "Authentication required"}, status=401)
             
        if batch.status not in ['Submitted', 'Manager Approved']:
            return Response({"error": "Batch is not in a valid status for approval"}, status=400)
            
        if batch.current_approver != user:
             return Response({"error": "Unauthorized: You are not the designated approver."}, status=403)

        # Sync any row-level rejections/remarks from frontend
        if 'data_json' in request.data and request.data['data_json']:
            batch.data_json = request.data['data_json']
            batch.save()

        is_hr = _is_hr(user)
        requester = batch.user

        # Count newly/previously rejected rows for better notification
        rejected_rows = [r for r in batch.data_json if r.get('_status') == 'Rejected']
        rejection_note = f" (with {len(rejected_rows)} lines rejected)" if rejected_rows else ""
        
        # If the approver is not HR, do management chain logic
        if not is_hr:
            next_approver = None
            
            # Try explicit levels first
            if batch.hierarchy_level == 1:
                next_approver = getattr(batch, 'senior_manager', None) or getattr(batch, 'hod_director', None)
            elif batch.hierarchy_level == 2:
                next_approver = getattr(batch, 'hod_director', None)
            
            # DYNAMIC FALLBACK: If no explicit level but current user has a manager
            if not next_approver:
                potential_manager = user.reporting_manager
                # Ensure we don't route back to requester or to a non-existent manager
                if potential_manager and potential_manager != user and potential_manager != requester:
                    next_approver = potential_manager

            if next_approver:
                # Move to next manager in chain
                batch.current_approver = next_approver
                batch.hierarchy_level += 1
                batch.save()
                
                if batch.trip:
                    batch.trip.current_approver = next_approver
                    batch.trip.hierarchy_level = batch.hierarchy_level
                    batch.trip.save()
                    update_trip_lifecycle(batch.trip, f"Level {batch.hierarchy_level - 1} Approval", f"Bulk travel log approved by {user.name} and forwarded to {next_approver.name}.")
                
                Notification.objects.create(
                    user=next_approver,
                    title="Pending Approval: Bulk Travel Log",
                    message=f"{requester.name}'s bulk travel log requires your review (Forwarded from {user.name}).",
                    type='info'
                )
                Notification.objects.create(
                    user=requester,
                    title=f"Forwarded by {user.name}",
                    message=f"Your bulk travel log has been reviewed by {user.name}{rejection_note} and forwarded to {next_approver.name}.",
                    type='warning' if rejected_rows else 'success'
                )
                return Response({"message": f"Batch approved and forwarded to {next_approver.name}."})
            else:
                # End of Management Chain -> Move to HR Approval
                hr_head = get_hr_head(requester)
                batch.status = 'Manager Approved'
                batch.current_approver = hr_head
                batch.save()
                
                if batch.trip:
                    batch.trip.status = 'Manager Approved'
                    batch.trip.current_approver = hr_head
                    batch.trip.save()
                    update_trip_lifecycle(batch.trip, f"Approved by {user.name}", f"Approved by {user.name} and forwarded to HR ({hr_head.name if hr_head else 'Head'}).")
                
                Notification.objects.create(
                    user=requester,
                    title="Management Reviewed",
                    message=f"Your bulk travel log has been reviewed by management{rejection_note} and sent to HR for verification.",
                    type='warning' if rejected_rows else 'success'
                )
                
                if hr_head:
                    Notification.objects.create(
                        user=hr_head,
                        title="HR Verification Required",
                        message=f"{requester.name}'s bulk travel log has been approved by {user.name} and awaits your verification.",
                        type='info'
                    )
                return Response({"message": "Sent to HR for verification"})

        # --- STAGE 2: HR Approval ---
        # If we reach here, it's the final approval (HR)
        # --- Final Approval Side Effects ---
        # Activate Trip Story
        if batch.trip:
            update_trip_lifecycle(batch.trip, "Trip Story Activated", "Bulk activity log approved. Trip story and claims are now active.")
            # If there's a status that signifies 'Ready for Claim', we could set it here.
            # Usually 'Approved' or 'Completed' works.
            if batch.trip.consider_as_local:
                batch.trip.status = 'Approved' # Active for local
                batch.trip.save()
        # Create Expenses (Activities) for each row only when the final manager approves
        created_ids = []
        for row in (batch.data_json or []):
            # SKIP REJECTED ROWS during final creation
            if row.get('_status') == 'Rejected':
                continue

            # 1. Get Origin and Destination from JSON
            origin = str(row.get('origin_route', '')).strip()
            destination = str(row.get('destination_route', '')).strip()

            # 2. Map mode and subType to match DynamicExpenseGrid's expected options
            # Since mode is not provided in the bulk template, we default to Car/Own Car
            excel_mode = str(row.get('mode', '')).lower()
            excel_vehicle = str(row.get('vehicle', '')).lower()
            
            mapped_mode = "Car / Cab"
            mapped_subType = "Own Car"
            
            if 'bike' in excel_mode or '2 wheeler' in excel_mode:
                mapped_mode = "Bike"
                mapped_subType = "Own Bike" if 'own' in excel_vehicle else "Ride Bike"
            elif 'bus' in excel_mode or 'metro' in excel_mode or 'public' in excel_mode:
                mapped_mode = "Public Transport"
                if 'metro' in excel_mode: mapped_subType = "Metro"
                elif 'bus' in excel_mode: mapped_subType = "Local Bus"
                else: mapped_subType = "Auto" # Default for PT
            elif excel_mode: # Only if provided
                if 'own' in excel_vehicle: mapped_subType = "Own Car"
                elif 'company' in excel_vehicle: mapped_subType = "Company Car"
                elif 'ride' in excel_vehicle or 'uber' in excel_vehicle or 'ola' in excel_vehicle or 'taxi' in excel_mode: 
                    mapped_subType = "Ride Hailing"

            # Pull and normalise the date. skip any bogus rows (e.g. the
            # instruction/sample line that sneaked in from earlier uploads).
            clean_date = str(row.get('date', '')).strip()
            if not clean_date or 'instruc' in clean_date.lower():
                # invalid/empty date – skip this row entirely
                continue
            # ensure format is YYYY-MM-DD; try parsing to detect bad values
            try:
                # this will raise ValueError for non‑ISO strings like "?  Instruc"
                datetime.date.fromisoformat(clean_date)
            except Exception:
                continue

            if len(str(clean_date)) > 10:
                clean_date = str(clean_date)[:10] # type: ignore

            desc_json = {
                "natureOfVisit": row.get('visit_intent'), # Updated to use visit_intent as purpose was removed
                "remarks": row.get('remarks'),
                "from_bulk_upload": True,
                "origin": origin,
                "destination": destination,
                "mode": mapped_mode,
                "subType": mapped_subType,
                "odoStart": row.get('odo_start', 0),
                "odoEnd": row.get('odo_end', 0),
                "time": {
                    "boardingDate": clean_date,
                    "boardingTime": row.get('start_time', '09:00'),
                    "actualTime": row.get('reach_time', '18:00')
                }
            }
            
            def safe_float(val):
                try: 
                    return float(val) if val not in [None, '', '-', 'nan', 'NaN'] else 0.0 # type: ignore
                except (ValueError, TypeError): 
                    return 0.0

            odo_s = safe_float(row.get('odo_start', 0))
            odo_e = safe_float(row.get('odo_end', 0))

            expense = Expense.objects.create(
                trip=batch.trip,
                date=clean_date,
                category='Fuel', 
                amount=0,       
                paid_by='Self (Out of Pocket)',
                description=json.dumps(desc_json),
                status='Approved', 
                odo_start=odo_s,
                odo_end=odo_e,
                distance=max(0.0, float(odo_e - odo_s)),
                travel_mode=mapped_mode,
                vehicle_type=mapped_subType,
                latitude=0, longitude=0
            )
            created_ids.append(expense.id)
            
        batch.status = 'Approved'
        batch.created_expenses = created_ids
        batch.save()
        
        # Notify user
        Notification.objects.create(
            user=batch.user,
            title="Bulk Activities Approved",
            message=f"Your travel log from {batch.file_name} has been approved and added to your report.",
            type='success'
        )
        
        return Response({"message": "Batch approved and activities created"})

    @action(detail=True, methods=['post'], url_path='reject')
    def reject_batch(self, request, pk=None):
        batch = self.get_object()
        user = getattr(request, 'custom_user', None)
        
        if batch.current_approver != user:
             return Response({"error": "Unauthorized"}, status=403)
             
        if 'data_json' in request.data:
            batch.data_json = request.data['data_json']
             
        batch.status = 'Rejected'
        batch.remarks = request.data.get('remarks', 'Rejected by Manager')
        batch.save()
        
        # Notify user
        Notification.objects.create(
            user=batch.user,
            title="Bulk Activities Rejected",
            message=f"Your travel log {batch.file_name} was rejected. Reason: {batch.remarks}",
            type='error'
        )
        return Response({"message": "Batch rejected"})

        return Response({"message": "Batch rejected"})

from django.db import models
from rest_framework.decorators import action

# --- MASTER VIEWSETS ---

class MasterActionMixin:
    """Mixin to provide restore and include_deleted functionality to Master ViewSets."""
    def get_queryset(self):
        include_deleted = self.request.query_params.get('include_deleted') == 'true'
        if include_deleted:
            return self.queryset.model.all_objects.all().order_by('id')
        return self.queryset.order_by('id')

    def perform_create(self, serializer):
        # Master unique name check: if deleted record exists, restore it instead
        model = self.queryset.model
        unique_field = None
        
        # Identify the unique character field (usually mode_name, class_name, etc.)
        for field in model._meta.fields:
            if isinstance(field, (models.CharField, models.TextField)) and field.unique:
                unique_field = field.name
                break
        
        if unique_field:
            val = serializer.validated_data.get(unique_field)
            if val:
                existing = model.all_objects.filter(**{unique_field: val}, is_deleted=True).first()
                if existing:
                    existing.is_deleted = False
                    existing.deleted_at = None
                    # Update other fields to the new values provided
                    for k, v in serializer.validated_data.items():
                        setattr(existing, k, v)
                    existing.save()
                    # Return the restored object
                    serializer.instance = existing
                    return
        
        serializer.save()

    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        instance = self.queryset.model.all_objects.get(pk=pk)
        instance.is_deleted = False
        instance.deleted_at = None
        instance.deleted_by = None
        instance.save()
        return Response({"status": "restored"})

class TravelModeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = TravelModeMaster.objects.all()
    serializer_class = TravelModeMasterSerializer

class BookingTypeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = BookingTypeMaster.objects.all()
    serializer_class = BookingTypeMasterSerializer

class OperatorMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = OperatorMaster.objects.all()
    serializer_class = OperatorMasterSerializer

class TravelClassMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = TravelClassMaster.objects.all()
    serializer_class = TravelClassMasterSerializer

class VehicleMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = VehicleMaster.objects.all()
    serializer_class = VehicleMasterSerializer

class ProviderMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = ProviderMaster.objects.all()
    serializer_class = ProviderMasterSerializer

class TicketStatusMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = TicketStatusMaster.objects.all()
    serializer_class = TicketStatusMasterSerializer

class QuotaTypeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = QuotaTypeMaster.objects.all()
    serializer_class = QuotaTypeMasterSerializer

class LocalTravelModeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = LocalTravelModeMaster.objects.all()
    serializer_class = LocalTravelModeMasterSerializer

class LocalProviderMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = LocalProviderMaster.objects.all()
    serializer_class = LocalProviderMasterSerializer

class LocalSubTypeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = LocalSubTypeMaster.objects.all()
    serializer_class = LocalSubTypeMasterSerializer

class StayTypeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = StayTypeMaster.objects.all()
    serializer_class = StayTypeMasterSerializer

class RoomTypeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = RoomTypeMaster.objects.all()
    serializer_class = RoomTypeMasterSerializer

class StayBookingTypeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = StayBookingTypeMaster.objects.all()
    serializer_class = StayBookingTypeMasterSerializer

class StayBookingSourceMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = StayBookingSourceMaster.objects.all()
    serializer_class = StayBookingSourceMasterSerializer

class MealCategoryMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = MealCategoryMaster.objects.all()
    serializer_class = MealCategoryMasterSerializer

class MealTypeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = MealTypeMaster.objects.all()
    serializer_class = MealTypeMasterSerializer

class MealSourceMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = MealSourceMaster.objects.all()
    serializer_class = MealSourceMasterSerializer

class MealProviderMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = MealProviderMaster.objects.all()
    serializer_class = MealProviderMasterSerializer

class JobReportViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = JobReport.objects.all()
    serializer_class = JobReportSerializer

    def get_queryset(self):
        trip_id = self.request.query_params.get('trip_id')
        if trip_id:
            return self.queryset.filter(trip_id=trip_id)
        return self.queryset

    def perform_create(self, serializer):
        serializer.save(user=self.request.custom_user)

class IncidentalTypeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = IncidentalTypeMaster.objects.all()
    serializer_class = IncidentalTypeMasterSerializer

class MasterModuleViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = MasterModule.objects.all()
    serializer_class = MasterModuleSerializer

    @action(detail=False, methods=['get'])
    @permission_classes_decorator([AllowAny])
    def seed_metadata(self, request):
        # 0. Clean Sweep (Ensure fresh definitions)
        CustomMasterDefinition.objects.all().delete()
        MasterModule.objects.all().delete()

        # 1. Modules
        modules_data = [
            {'name': 'Travel', 'display_order': 1},
            {'name': 'Local Conveyance', 'display_order': 2},
            {'name': 'Stay', 'display_order': 3},
            {'name': 'Food', 'display_order': 4},
            {'name': 'Incidental', 'display_order': 5},
        ]
        
        modules = {}
        for m_data in modules_data:
            mod, _ = MasterModule.objects.update_or_create(
                name=m_data['name'],
                defaults={'display_order': m_data['display_order']}
            )
            modules[m_data['name']] = mod

        # 2. Definitions
        definitions_data = [
            {'module': 'Travel', 'table_name': 'Travel Modes', 'endpoint': 'travel-mode-masters', 'fields': 'mode_name,status'},
            {'module': 'Travel', 'table_name': 'Booking Types', 'endpoint': 'booking-type-masters', 'fields': 'booking_type,status'},
            {'module': 'Travel', 'table_name': 'Operators (Flight/Train/Bus)', 'endpoint': 'operator-masters', 'fields': 'operator_name,is_flight,is_train,is_bus,status'},
            {'module': 'Travel', 'table_name': 'Travel Classes', 'endpoint': 'travel-class-masters', 'fields': 'class_name,is_flight,is_train,is_bus,status'},
            {'module': 'Travel', 'table_name': 'Vehicles', 'endpoint': 'vehicle-masters', 'fields': 'vehicle_name,is_bus,is_intercity_cab,status'},
            {'module': 'Travel', 'table_name': 'Providers', 'endpoint': 'provider-masters', 'fields': 'provider_name,is_flight,is_train,is_bus,is_intercity_cab,status'},
            {'module': 'Travel', 'table_name': 'Ticket Statuses', 'endpoint': 'ticket-status-masters', 'fields': 'status_name,is_flight,is_train,is_bus,is_intercity_cab,status'},
            {'module': 'Travel', 'table_name': 'Quota Types', 'endpoint': 'quota-type-masters', 'fields': 'quota_name,status'},
            
            {'module': 'Local Conveyance', 'table_name': 'Local Modes', 'endpoint': 'local-travel-mode-masters', 'fields': 'mode_name,status'},
            {'module': 'Local Conveyance', 'table_name': 'Local Providers', 'endpoint': 'local-provider-masters', 'fields': 'provider_name,is_car,is_bike,is_auto,is_bus,is_metro,status'},
            {'module': 'Local Conveyance', 'table_name': 'Local Sub-Types', 'endpoint': 'local-sub-type-masters', 'fields': 'sub_type,is_car,is_bike,is_auto,status'},
            
            {'module': 'Stay', 'table_name': 'Stay Types', 'endpoint': 'stay-type-masters', 'fields': 'stay_type,status'},
            {'module': 'Stay', 'table_name': 'Room Types', 'endpoint': 'room-type-masters', 'fields': 'room_type,status'},
            {'module': 'Stay', 'table_name': 'Stay Booking Types', 'endpoint': 'stay-booking-type-masters', 'fields': 'booking_type,status'},
            {'module': 'Stay', 'table_name': 'Stay Booking Sources', 'endpoint': 'stay-booking-source-masters', 'fields': 'source_name,status'},
            
            {'module': 'Food', 'table_name': 'Meal Categories', 'endpoint': 'meal-category-masters', 'fields': 'category_name,status'},
            {'module': 'Food', 'table_name': 'Meal Types', 'endpoint': 'meal-type-masters', 'fields': 'meal_type,status'},
            {'module': 'Food', 'table_name': 'Meal Sources', 'endpoint': 'meal-source-masters', 'fields': 'source_name,status'},
            {'module': 'Food', 'table_name': 'Meal Providers', 'endpoint': 'meal-provider-masters', 'fields': 'provider_name,status'},
            
            {'module': 'Incidental', 'table_name': 'Incidental Types', 'endpoint': 'incidental-type-masters', 'fields': 'expense_type,category,status'},
        ]
        
        for d_data in definitions_data:
            CustomMasterDefinition.objects.update_or_create(
                table_name=d_data['table_name'],
                defaults={
                    'module_ref': modules[d_data['module']],
                    'api_endpoint': d_data['endpoint'],
                    'fields_list': d_data['fields'],
                    'is_system': True
                }
            )

        # 3. Seed Initial Records for popular masters
        # Stay Booking Types
        for btype in ['Self Booking', 'Company Booking', 'Online Booking']:
            StayBookingTypeMaster.objects.get_or_create(booking_type=btype, defaults={'status': True})
        
        # Room Types
        for rtype in ['Standard', 'Deluxe', 'Executive', 'Suite']:
            RoomTypeMaster.objects.get_or_create(room_type=rtype, defaults={'status': True})

        # Stay Types
        for stype in ['Hotel Stay', 'Guest House', 'Self Stay']:
            StayTypeMaster.objects.get_or_create(stay_type=stype, defaults={'status': True})

        # Local Modes
        for lmode in ['Car', 'Bike', 'Auto', 'Public Transport']:
            LocalTravelModeMaster.objects.get_or_create(mode_name=lmode, defaults={'status': True})

        # Meal Categories
        for mcat in ['Self Meal', 'Business Lunch', 'Dinner with Client']:
            MealCategoryMaster.objects.get_or_create(category_name=mcat, defaults={'status': True})

        return Response({
            "status": "Master Metadata and Initial Records Seeded Successfully", 
            "modules": len(modules_data), 
            "definitions": len(definitions_data),
            "records_note": "Added initial values for Stay Booking Types, Room Types, Stay Types, Local Modes, and Meal Categories."
        })

class CustomMasterDefinitionViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = CustomMasterDefinition.objects.all()
    serializer_class = CustomMasterDefinitionSerializer

class CustomMasterValueViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = CustomMasterValue.objects.all()
    serializer_class = CustomMasterValueSerializer
    filterset_fields = ['definition']


