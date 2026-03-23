"""
Travel Expense Statement Export Views
Generates PDF and Excel reports matching the BAVYA Travel Expenses Statement template.
"""

import os
import io
import json

from django.http import HttpResponse
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from core.permissions import IsCustomAuthenticated
from .models import Trip

# ── ReportLab ────────────────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image
)

# ── OpenPyXL ─────────────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette ───────────────────────────────────────────────────────────
HDR_BLUE   = colors.HexColor('#1e3a5f')
YELLOW_HL  = colors.HexColor('#FFD700')
LIGHT_GRAY = colors.HexColor('#f5f5f5')
WHITE      = colors.white


# ═════════════════════════════════════════════════════════════════════════════
#  Data builder  –  reads real Expense objects from the DB
# ═════════════════════════════════════════════════════════════════════════════
# Intercity travel categories (go to Section I)
INTERCITY_CATEGORIES = {'Fuel', 'Travel'}
# "Other particulars" categories (go to Section II right column)
OTHER_CATEGORIES = {'Food', 'Accommodation', 'Incidental', 'Others'}
# Local conveyance categories (go to Section II left table)
LOCAL_CATEGORIES = {'Toll'}


def _parse_desc(raw: str) -> dict:
    """Try to JSON-decode a description field; return {} on failure."""
    if raw and raw.strip().startswith('{'):
        try:
            return json.loads(raw)
        except Exception:
            pass
    return {}

def format_inr(amount) -> str:
    """Formats a number in Indian currency style (e.g. 12,34,567.89)."""
    try:
        amount = float(amount or 0)
        is_negative = amount < 0
        amount = abs(amount)
        s, *d = str(round(amount, 2)).split('.')
        d = d[0].ljust(2, '0') if d else '00'
        res = s[-3:]
        s = s[:-3]
        while s:
            res = s[-2:] + ',' + res
            s = s[:-2]
        formatted = f"{res}.{d}"
        return f"-{formatted}" if is_negative else formatted
    except Exception:
        return '0.00'


def _build_statement_data(trip: Trip) -> dict:
    user = trip.user
    emp_name    = trip.user_name or (user.name if user and user.name else 'N/A')
    emp_code    = (user.employee_id if user and user.employee_id else 'N/A') if user else 'N/A'
    project     = trip.project_code or 'General'
    # User requested Base Location to be the trip's starting location (source)
    base_loc    = trip.source or 'N/A'
    month_label = trip.start_date.strftime('%B %Y') if trip.start_date else 'N/A'

    # ── Walk through every expense ──────────────────────────────────────────
    travel_rows  = []   # Section I
    local_rows   = []   # Section II – left
    other_totals = {'Food': 0.0, 'Accommodation': 0.0, 'Incidental': 0.0, 'Others': 0.0}

    travel_fare  = 0.0
    lodging      = 0.0
    daily_allow  = 0.0
    incidental   = 0.0
    local_conv   = 0.0
    other_misc   = 0.0
    food_total   = 0.0

    expenses = list(trip.expenses.all().order_by('date'))

    for exp in expenses:
        amt  = float(exp.amount or 0)
        desc = _parse_desc(exp.description or '')

        # Prefer JSON-embedded fields, fall back to model fields / trip fields
        origin      = desc.get('origin', '') or desc.get('from', '') or trip.source or ''
        destination = desc.get('destination', '') or desc.get('to', '') or trip.destination or ''
        mode        = desc.get('mode', '') or exp.travel_mode or exp.category or ''
        dep_time    = desc.get('dep_time', '') or desc.get('departure_time', '') or ''
        arr_time    = desc.get('arr_time', '') or desc.get('arrival_time', '') or ''
        arr_date    = desc.get('arr_date', '') or str(exp.date)
        dep_date    = desc.get('dep_date', '') or str(exp.date)
        km          = float(exp.distance or desc.get('distance', 0) or 0)

        cat = exp.category or ''

        if cat in INTERCITY_CATEGORIES or (exp.travel_mode and exp.travel_mode.strip()):
            # Long-distance / intercity → Section I
            travel_fare += amt
            travel_rows.append({
                'dep_date': dep_date,
                'dep_time': dep_time,
                'from': origin,
                'arr_date': arr_date,
                'arr_time': arr_time,
                'to': destination,
                'mode': mode,
                'km': km,
                'amount': amt,
            })

        elif cat == 'Food':
            food_total += amt
            other_totals['Food'] += amt

        elif cat == 'Accommodation':
            lodging += amt
            other_totals['Accommodation'] += amt

        elif cat == 'Incidental':
            incidental += amt
            other_totals['Incidental'] += amt

        elif cat == 'Others':
            other_misc += amt
            other_totals['Others'] += amt

        else:
            # Toll, local car, etc. → Section II left
            local_conv += amt
            local_rows.append({
                'date': str(exp.date),
                'from': origin,
                'to': destination,
                'mode': mode,
                'amount': amt,
            })

    # Section II right column – show all 4 "Other" buckets (always present, real amounts)
    local_other = [
        ('Food:',          other_totals['Food']),
        ('Accommodation:', other_totals['Accommodation']),
        ('Incidental:',    other_totals['Incidental']),
        ('Miscellaneous:', other_totals['Others']),
    ]

    grand_total = travel_fare + lodging + incidental + local_conv + other_misc + food_total

    # Advances that are actually released
    advances = float(sum(
        float(a.executive_approved_amount or a.requested_amount or 0)
        for a in trip.advances.filter(status__in=['COMPLETED', 'Paid', 'Transferred', 'COMPLETED'])
    ))
    to_be_refunded   = max(0.0, advances - grand_total)
    to_be_reimbursed = max(0.0, grand_total - advances)

    return {
        'emp_name':   emp_name,
        'emp_code':   emp_code,
        'project':    project,
        'base_loc':   base_loc,
        'month_label': month_label,
        'bank_name':  (user.bank_name if user and user.bank_name else 'N/A'),
        'account_no': (user.account_no if user and user.account_no else 'N/A'),
        'ifsc_code':  (user.ifsc_code if user and user.ifsc_code else 'N/A'),
        'travel_rows': travel_rows,
        'local_rows':  local_rows,
        'local_other': local_other,
        'summary': {
            'Total Travel Fare':    travel_fare,
            'Total Lodging':        lodging,
            'Daily Allowance':      food_total, # Map food_total to Daily Allowance
            'Incidental Expenses':  incidental,
            'Local Conveyance':     local_conv,
            'Other Misc.':          other_misc,
            'Own Stay Allowance':   0.0,
        },
        'grand_total':      grand_total,
        'advance_taken':    advances,
        'to_be_refunded':   to_be_refunded,
        'to_be_reimbursed': to_be_reimbursed,
        'trip_id': trip.trip_id,
        'members': ", ".join(trip.members) if isinstance(trip.members, list) else (trip.members or ""),
    }


# ═════════════════════════════════════════════════════════════════════════════
#  PDF Generator
# ═════════════════════════════════════════════════════════════════════════════
def generate_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    PAGE = landscape(A4)
    MARGIN = 10 * mm
    W = PAGE[0] - 2 * MARGIN          # usable width ≈ 277 mm

    doc = SimpleDocTemplate(
        buf, pagesize=PAGE,
        rightMargin=MARGIN, leftMargin=MARGIN,
        topMargin=MARGIN, bottomMargin=MARGIN,
    )

    SS = getSampleStyleSheet()

    def ps(name, **kw):
        return ParagraphStyle(name, parent=SS['Normal'], **kw)

    # Shared styles
    S_title  = ps('T',  fontSize=13, fontName='Helvetica-Bold', alignment=TA_CENTER)
    S_label  = ps('L',  fontSize=8.5, fontName='Helvetica-Bold')
    S_val    = ps('V',  fontSize=8.5, fontName='Helvetica')
    S_hdr    = ps('H',  fontSize=8,  fontName='Helvetica-Bold',
                  textColor=WHITE, alignment=TA_CENTER, leading=11)
    S_cell   = ps('C',  fontSize=8,  fontName='Helvetica', leading=10)
    S_cellR  = ps('CR', fontSize=8,  fontName='Helvetica', leading=10, alignment=TA_RIGHT)
    S_sec    = ps('SC', fontSize=9.5, fontName='Helvetica-Bold',
                  textColor=colors.HexColor('#1e3a5f'))
    S_smLbl  = ps('SL', fontSize=8,  fontName='Helvetica')
    S_smBld  = ps('SB', fontSize=8,  fontName='Helvetica-Bold')
    S_smR    = ps('SR', fontSize=8,  fontName='Helvetica', alignment=TA_RIGHT)
    S_smBR   = ps('SBR',fontSize=8,  fontName='Helvetica-Bold', alignment=TA_RIGHT)

    def row_bg(idx):               # alternating row colour (idx starts at 0)
        return LIGHT_GRAY if idx % 2 == 1 else WHITE

    elements = []

    # ── Logo + Title ─────────────────────────────────────────────────────────
    # Logo Path (Try multiple locations)
    logo_path = None
    possible_paths = [
        # Relative to backend root
        os.path.join(settings.BASE_DIR, '..', 'TGS FRONTEND', 'public', 'bavya.png'),
        # Hardcoded absolute (Windows)
        r'e:\TGS-V11\TGS-V1\TGS FRONTEND\public\bavya.png',
    ]
    for p in possible_paths:
        if os.path.exists(p):
            logo_path = p
            break

    if logo_path:
        logo_img = Image(logo_path, width=35*mm, height=12*mm)
    else:
        # Fallback text if logo missing
        logo_img = Paragraph('<b><font color="#e74c3c">■</font> BAVYA</b>',
                            ps('logo', fontSize=14, fontName='Helvetica-Bold'))
    
    title_p = Paragraph(
        f'Travel Expenses Statement for the month of {data["month_label"]}', S_title)

    ht = Table([[logo_img, title_p, '']], colWidths=[45*mm, W - 90*mm, 45*mm])
    ht.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements += [ht, Spacer(1, 2*mm),
                 HRFlowable(width=W, thickness=0.8, color=colors.HexColor('#bbbbbb')),
                 Spacer(1, 2*mm)]

    # ── Employee / Bank info ──────────────────────────────────────────────────
    info_data = [
        [S_label, Paragraph('<b>Employee Name:</b>', S_label),
         Paragraph(data['emp_name'], S_val),   '',
         Paragraph('<b>Bank Name:</b>',  S_label), Paragraph(data['bank_name'], S_val)],
        ['', Paragraph('<b>Employee Code:</b>', S_label),
         Paragraph(data['emp_code'], S_val),   '',
         Paragraph('<b>Account No:</b>', S_label), Paragraph(data['account_no'], S_val)],
        ['', Paragraph('<b>Project Name:</b>', S_label),
         Paragraph(data['project'], S_val),    '',
         Paragraph('<b>IFS Code:</b>',  S_label), Paragraph(data['ifsc_code'], S_val)],
        ['', Paragraph('<b>Trip Source:</b>', S_label),
         Paragraph(data['base_loc'], S_val),   '', 
         Paragraph('<b>Team Members:</b>', S_label), Paragraph(data['members'], S_val)],
    ]
    # fix: first col is just padding
    info_data = [row[1:] for row in info_data]   # drop the phantom first cell
    cw_info = [38*mm, 55*mm, 12*mm, 32*mm, 55*mm]
    ti = Table(info_data, colWidths=cw_info)
    ti.setStyle(TableStyle([
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING',    (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
    ]))
    elements += [ti, Spacer(1, 3*mm),
                 HRFlowable(width=W, thickness=0.5, color=colors.HexColor('#cccccc')),
                 Spacer(1, 2*mm)]

    # ─────────────────────────────────────────────────────────────────────────
    # Section I  –  Travel Particulars
    # ─────────────────────────────────────────────────────────────────────────
    elements.append(Paragraph('I. Travel Particulars', S_sec))
    elements.append(Spacer(1, 2*mm))

    # Column widths (must sum to W)
    cw_t = [7*mm, 22*mm, 13*mm, 32*mm, 22*mm, 13*mm, 32*mm, 32*mm, 14*mm, 14*mm, 22*mm]
    # sum = 223 mm  → ≈ W

    def travel_hdr():
        labels = ['Sl', 'Dep Date', 'Time', 'From', 'Arr Date', 'Time', 'To', 'Mode', 'KM', 'Rate', 'Amount']
        return [Paragraph(l, S_hdr) for l in labels]

    t1_rows = [travel_hdr()]
    for i, r in enumerate(data['travel_rows']):
        t1_rows.append([
            Paragraph(str(i + 1), S_cell),
            Paragraph(r['dep_date'], S_cell),
            Paragraph(r['dep_time'], S_cell),
            Paragraph(r['from'],     S_cell),
            Paragraph(r['arr_date'], S_cell),
            Paragraph(r['arr_time'], S_cell),
            Paragraph(r['to'],       S_cell),
            Paragraph(r['mode'],     S_cell),
            Paragraph(f"{r['km']:.0f}" if r['km'] else '', S_cellR),
            Paragraph('', S_cell),
            Paragraph(format_inr(r['amount']), S_cellR),
        ])

    # No padding rows - strict fit to data

    s1_style = TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), HDR_BLUE),
        ('TEXTCOLOR',     (0, 0), (-1, 0), WHITE),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, -1), 8),
        ('ALIGN',         (0, 0), (-1, 0),  'CENTER'),
        ('ALIGN',         (8, 1), (10, -1), 'RIGHT'),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#aaaaaa')),
        ('TOPPADDING',    (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
    ])
    for i in range(1, len(t1_rows)):
        if i % 2 == 0:
            s1_style.add('BACKGROUND', (0, i), (-1, i), LIGHT_GRAY)

    tt = Table(t1_rows, colWidths=cw_t, repeatRows=1)
    tt.setStyle(s1_style)
    elements += [tt, Spacer(1, 4*mm)]

    # ─────────────────────────────────────────────────────────────────────────
    # Section II  –  Local Conveyance  +  Summary (side by side)
    # ─────────────────────────────────────────────────────────────────────────
    elements.append(Paragraph('II. Local Conveyance &amp; Other Expenses', S_sec))
    elements.append(Spacer(1, 2*mm))

    # --- Left sub-table: local conveyance + "Particulars (Other)" ---
    # columns: Sl | Date | From | To | Mode | Amt || Particulars | Amt
    cw_l = [7*mm, 20*mm, 24*mm, 24*mm, 22*mm, 18*mm, 38*mm, 18*mm]
    L_W  = sum(cw_l)    # 171 mm

    def local_hdr():
        labels = ['Sl', 'Date', 'From', 'To', 'Mode', 'Amt', 'Particulars (Other)', 'Amt']
        return [Paragraph(l, S_hdr) for l in labels]

    n_rows = max(len(data['local_rows']), len(data['local_other']))
    l2_rows = [local_hdr()]
    for i in range(n_rows):
        lr = data['local_rows'][i] if i < len(data['local_rows']) else {}
        op = data['local_other'][i] if i < len(data['local_other']) else ('', 0.0)
        l2_rows.append([
            Paragraph(str(i + 1) if lr else '', S_cell),
            Paragraph(lr.get('date', ''),       S_cell),
            Paragraph(lr.get('from', ''),       S_cell),
            Paragraph(lr.get('to',   ''),       S_cell),
            Paragraph(lr.get('mode', ''),       S_cell),
            Paragraph(format_inr(lr['amount']) if lr.get('amount') else '', S_cellR),
            Paragraph(op[0],                    S_cell),
            Paragraph(format_inr(op[1]) if op[1] else '', S_cellR),
        ])

    s2_lft = TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), HDR_BLUE),
        ('TEXTCOLOR',     (0, 0), (-1, 0), WHITE),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, -1), 8),
        ('ALIGN',         (0, 0), (-1, 0),  'CENTER'),
        ('ALIGN',         (5, 1), (5, -1),  'RIGHT'),
        ('ALIGN',         (7, 1), (7, -1),  'RIGHT'),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#aaaaaa')),
        ('TOPPADDING',    (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
    ])
    for i in range(1, len(l2_rows)):
        if i % 2 == 0:
            s2_lft.add('BACKGROUND', (0, i), (-1, i), LIGHT_GRAY)

    tl = Table(l2_rows, colWidths=cw_l, repeatRows=1)
    tl.setStyle(s2_lft)

    # --- Right sub-table: Summary ---
    SUM_W   = W - L_W - 8*mm          # gap = 8 mm
    SUM_LW  = SUM_W - 30*mm
    SUM_RW  = 30*mm
    cw_s    = [SUM_LW, SUM_RW]

    sum_rows_data = list(data['summary'].items())
    grand_idx = len(sum_rows_data) + 1     # +1 for header row

    s_data = [[Paragraph('<b>Summary</b>', S_smBld),
               Paragraph('<b>Amount (INR)</b>', S_smBR)]]
    for lbl, val in sum_rows_data:
        s_data.append([Paragraph(lbl, S_smLbl),
                        Paragraph(format_inr(val), S_smR)])
    # Grand total (yellow)
    s_data.append([Paragraph('<b>Grand Total</b>', S_smBld),
                   Paragraph(f"<b>{format_inr(data['grand_total'])}</b>", S_smBR)])
    # Extra rows
    for lbl, val in [
        ('Advance Taken',    data['advance_taken']),
        ('To be Refunded',   data['to_be_refunded']),
        ('To be Reimbursed', data['to_be_reimbursed']),
    ]:
        s_data.append([Paragraph(lbl, S_smLbl),
                        Paragraph(format_inr(val), S_smR)])

    s2_rgt = TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#d0d0d0')),
        ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, -1), 8),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.HexColor('#aaaaaa')),
        ('BACKGROUND',    (0, grand_idx), (-1, grand_idx), YELLOW_HL),
        ('TOPPADDING',    (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN',         (1, 0), (1, -1),  'RIGHT'),
    ])
    for i in range(1, grand_idx):
        if i % 2 == 0:
            s2_rgt.add('BACKGROUND', (0, i), (-1, i), LIGHT_GRAY)

    ts = Table(s_data, colWidths=cw_s)
    ts.setStyle(s2_rgt)

    # Side-by-side container
    outer = Table([[tl, Spacer(8*mm, 1), ts]],
                  colWidths=[L_W, 8*mm, SUM_W])
    outer.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    elements.append(outer)

    # ─────────────────────────────────────────────────────────────────────────
    # Signature Block
    # ─────────────────────────────────────────────────────────────────────────
    elements.append(Spacer(1, 8*mm))

    sig_line  = '_' * 22
    sig_style = ps('sig', fontSize=8.5, fontName='Helvetica-Bold',
                   alignment=TA_CENTER, spaceAfter=2)
    sig_line_style = ps('sigL', fontSize=9, fontName='Helvetica',
                        alignment=TA_CENTER, textColor=colors.HexColor('#333333'))

    SIG_LABELS = ['Signature - Employee', 'Reporting Authority', 'HR', 'Accounts']
    sig_col_w  = [W / 4] * 4

    sig_lines_row = [Paragraph(sig_line, sig_line_style) for _ in SIG_LABELS]
    sig_names_row = [Paragraph(f'<b>{lbl}</b>', sig_style) for lbl in SIG_LABELS]

    sig_table = Table([sig_lines_row, sig_names_row], colWidths=sig_col_w)
    sig_table.setStyle(TableStyle([
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN',        (0, 0), (-1, -1), 'BOTTOM'),
        ('TOPPADDING',    (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 4),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 2),
    ]))
    elements.append(sig_table)

    # ─────────────────────────────────────────────────────────────────────────
    # Footer  –  Generated on … | Trip ID: …
    # ─────────────────────────────────────────────────────────────────────────
    from datetime import datetime as _dt
    generated_at = _dt.now().strftime('%d/%m/%Y, %I:%M %p')
    footer_p = Paragraph(
        f'<i>Generated on {generated_at} &nbsp;|&nbsp; Trip ID: {data["trip_id"]}</i>',
        ps('footer', fontSize=7.5, fontName='Helvetica',
           alignment=TA_CENTER, textColor=colors.HexColor('#666666')),
    )
    elements += [Spacer(1, 4*mm), footer_p]

    doc.build(elements)
    return buf.getvalue()


# ═════════════════════════════════════════════════════════════════════════════
#  Excel Generator
# ═════════════════════════════════════════════════════════════════════════════
def generate_excel(data: dict) -> bytes:
    # Template Path
    tpl_path = os.path.join(settings.BASE_DIR, '..', 'TGS FRONTEND', 'public', 'Travel Expense Statement 27022026.xlsx')
    if not os.path.exists(tpl_path):
        # Hardcoded absolute fallback
        tpl_path = r'e:\TGS-V11\TGS-V1\TGS FRONTEND\public\Travel Expense Statement 27022026.xlsx'

    if os.path.exists(tpl_path):
        wb = openpyxl.load_workbook(tpl_path)
    else:
        wb = openpyxl.Workbook()
    
    ws = wb.active
    ws.title = 'Expense Statement'

    # Logo from bavya.png
    logo_path = os.path.join(settings.BASE_DIR, '..', 'TGS FRONTEND', 'public', 'bavya.png')
    if not os.path.exists(logo_path):
        logo_path = r'e:\TGS-V11\TGS-V1\TGS FRONTEND\public\bavya.png'
    
    if os.path.exists(logo_path):
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(logo_path)
        img.width = 120  # Approx scaled
        img.height = 40
        ws.add_image(img, 'A1')

    # Fills
    BLUE_F   = PatternFill('solid', fgColor='1E3A5F')
    YELLOW_F = PatternFill('solid', fgColor='FFD700')
    GRAY_F   = PatternFill('solid', fgColor='D0D0D0')
    ALT_F    = PatternFill('solid', fgColor='F5F5F5')
    WHITE_F  = PatternFill('solid', fgColor='FFFFFF')

    thin   = Side(style='thin', color='BBBBBB')
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hf(): return Font(bold=True, color='FFFFFF', size=9)
    def bf(): return Font(bold=True,  size=9)
    def nf(): return Font(size=9)

    def wcell(ws, r, c, val, bold=False, fill=WHITE_F, halign='left', num_fmt=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(bold=bold, size=9)
        cell.fill = fill
        cell.border = bdr
        cell.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=True)
        if num_fmt:
            cell.number_format = num_fmt
        return cell

    def ycell(ws, r, c, val, bold=False):
        cell = ws.cell(row=r, column=c, value=val)
        cell.fill = YELLOW_F
        cell.font = Font(bold=bold, size=9)
        cell.border = bdr
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.number_format = '##,##,##0.00'
        return cell

    # ── Row 1: Title ────────────────────────────────────────────────────────
    # If using template, A1 might already have logo/title. 
    # Let's ensure title is set correctly if it's merged.
    ws['C1'] = f'Travel Expenses Statement for the month of {data["month_label"]}'
    ws['C1'].font = Font(bold=True, size=13, color='1E3A5F')
    ws['C1'].alignment = Alignment(horizontal='center', vertical='center')

    # ── Rows 2-5: Employee / Bank info ───────────────────────────────────────
    info = [
        ('Employee Name:', data['emp_name'], 'Bank Name:',   data['bank_name']),
        ('Employee Code:', data['emp_code'], 'Account No:',  data['account_no']),
        ('Project Name:',  data['project'],  'IFS Code:',    data['ifsc_code']),
        ('Trip Source:',   data['base_loc'], 'Team Members:', data['members']),
    ]
    for offset, (l1, v1, l2, v2) in enumerate(info, start=2):
        ws.cell(row=offset, column=1, value=l1).font = bf()
        ws.cell(row=offset, column=2, value=v1).font = nf()
        # Merge if not already merged in template
        try: ws.merge_cells(start_row=offset, start_column=2, end_row=offset, end_column=5)
        except: pass
        
        ws.cell(row=offset, column=7, value=l2).font = bf()
        ws.cell(row=offset, column=8, value=v2).font = nf()
        try: ws.merge_cells(start_row=offset, start_column=8, end_row=offset, end_column=10)
        except: pass
        ws.row_dimensions[offset].height = 16

    # ── Section I ────────────────────────────────────────────────────────────
    S1 = 7
    try: ws.merge_cells(f'A{S1}:K{S1}')
    except: pass
    c = ws.cell(row=S1, column=1, value='I. Travel Particulars')
    c.font = Font(bold=True, size=10, color='1E3A5F')
    ws.row_dimensions[S1].height = 18

    hdrs1 = ['Sl', 'Dep Date', 'Time', 'From', 'Arr Date', 'Time', 'To', 'Mode', 'KM', 'Rate', 'Amount']
    for ci, h in enumerate(hdrs1, 1):
        cell = ws.cell(row=S1 + 1, column=ci, value=h)
        cell.font = hf(); cell.fill = BLUE_F; cell.border = bdr
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[S1 + 1].height = 18

    s1_data_start = S1 + 2
    for i, r in enumerate(data['travel_rows']):
        ri   = s1_data_start + i
        fill = ALT_F if i % 2 == 1 else WHITE_F
        vals = [i+1, r['dep_date'], r['dep_time'], r['from'],
                r['arr_date'], r['arr_time'], r['to'], r['mode'],
                r['km'] or '', '', r['amount']]
        for ci, v in enumerate(vals, 1):
            wcell(ws, ri, ci, v, fill=fill,
                  halign='right' if ci in (9, 11) else 'left',
                  num_fmt='##,##,##0.00' if ci == 11 else None)
        ws.row_dimensions[ri].height = 16

    # No padding rows for Section I

    # ── Section II ───────────────────────────────────────────────────────────
    S2 = s1_data_start + len(data['travel_rows']) + 1
    try: ws.merge_cells(f'A{S2}:H{S2}')
    except: pass
    c = ws.cell(row=S2, column=1, value='II. Local Conveyance & Other Expenses')
    c.font = Font(bold=True, size=10, color='1E3A5F')
    ws.row_dimensions[S2].height = 18

    hdrs2a = ['Sl', 'Date', 'From', 'To', 'Mode', 'Amt']
    hdrs2b = ['Particulars (Other)', 'Amt']

    for ci, h in enumerate(hdrs2a, 1):
        cell = ws.cell(row=S2 + 1, column=ci, value=h)
        cell.font = hf(); cell.fill = BLUE_F; cell.border = bdr
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for ci, h in enumerate(hdrs2b, 7):
        cell = ws.cell(row=S2 + 1, column=ci, value=h)
        cell.font = hf(); cell.fill = BLUE_F; cell.border = bdr
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[S2 + 1].height = 18

    max_s2 = max(len(data['local_rows']), len(data['local_other']))
    for i in range(max_s2):
        ri   = S2 + 2 + i
        lr   = data['local_rows'][i]   if i < len(data['local_rows'])   else {}
        op   = data['local_other'][i]  if i < len(data['local_other'])  else ('', 0.0)
        fill = ALT_F if i % 2 == 1 else WHITE_F

        vals_a = [i+1 if lr else '', lr.get('date',''), lr.get('from',''),
                  lr.get('to',''), lr.get('mode',''), lr.get('amount','')]
        for ci, v in enumerate(vals_a, 1):
            wcell(ws, ri, ci, v, fill=fill,
                  halign='right' if ci == 6 else 'left',
                  num_fmt='##,##,##0.00' if ci == 6 and v else None)

        wcell(ws, ri, 7, op[0], fill=fill)
        wcell(ws, ri, 8, op[1] if op[1] else '', fill=fill, halign='right',
              num_fmt='##,##,##0.00' if op[1] else None)
        ws.row_dimensions[ri].height = 16

    # ── Summary block (columns 10-11) ────────────────────────────────────────
    SC = 10    # start column for summary
    wcell(ws, S2, SC,     'Summary',       bold=True, fill=GRAY_F, halign='center')
    wcell(ws, S2, SC + 1, 'Amount (INR)',  bold=True, fill=GRAY_F, halign='center')

    for off, (lbl, val) in enumerate(data['summary'].items(), 1):
        r    = S2 + off
        fill = ALT_F if off % 2 == 1 else WHITE_F
        wcell(ws, r, SC,     lbl, fill=fill)
        wcell(ws, r, SC + 1, val, fill=fill, halign='right', num_fmt='##,##,##0.00')

    gt_r = S2 + len(data['summary']) + 1
    ycell(ws, gt_r, SC,     'Grand Total',         bold=True)
    ycell(ws, gt_r, SC + 1, data['grand_total'],   bold=True)

    for off, (lbl, val) in enumerate([
        ('Advance Taken',    data['advance_taken']),
        ('To be Refunded',   data['to_be_refunded']),
        ('To be Reimbursed', data['to_be_reimbursed']),
    ], 1):
        r = gt_r + off
        wcell(ws, r, SC,     lbl)
        wcell(ws, r, SC + 1, val, halign='right', num_fmt='##,##,##0.00')

    # ── Column widths ─────────────────────────────────────────────────────────
    col_w = {1:4, 2:12, 3:8, 4:16, 5:12, 6:8, 7:16, 8:18, 9:8, 10:22, 11:14, 12:2, 13:2}
    for col, w in col_w.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = 'B2'
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ═════════════════════════════════════════════════════════════════════════════
#  API Views
# ═════════════════════════════════════════════════════════════════════════════
class ExpenseStatementPDFView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request, trip_id):
        try:
            trip = Trip.objects.select_related('user').prefetch_related(
                'expenses', 'advances', 'claim'
            ).get(trip_id=trip_id)
        except Trip.DoesNotExist:
            return Response({'error': 'Trip not found'}, status=status.HTTP_404_NOT_FOUND)

        user = request.custom_user
        role_name = (user.role.name if user.role else '').lower()
        if trip.user != user and role_name not in ['admin', 'finance', 'hr', 'cfo']:
            return Response({'error': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)

        data = _build_statement_data(trip)
        pdf_bytes = generate_pdf(data)

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="expense_statement_{trip_id}.pdf"'
        )
        return response


class ExpenseStatementExcelView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request, trip_id):
        try:
            trip = Trip.objects.select_related('user').prefetch_related(
                'expenses', 'advances', 'claim'
            ).get(trip_id=trip_id)
        except Trip.DoesNotExist:
            return Response({'error': 'Trip not found'}, status=status.HTTP_404_NOT_FOUND)

        user = request.custom_user
        role_name = (user.role.name if user.role else '').lower()
        if trip.user != user and role_name not in ['admin', 'finance', 'hr', 'cfo']:
            return Response({'error': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)

        data = _build_statement_data(trip)
        xlsx_bytes = generate_excel(data)

        response = HttpResponse(
            xlsx_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="expense_statement_{trip_id}.xlsx"'
        )
        return response
